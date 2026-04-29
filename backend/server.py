import os
import logging
import uuid
import re
import random
import string
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, List
from dotenv import load_dotenv

import base64 as _b64
from fastapi import FastAPI, APIRouter, Depends, HTTPException, Request, Response, Query
from fastapi.responses import Response as FastResponse
from starlette.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr, Field
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_, desc

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from database import get_db, engine, Base
import models as M
from auth import (
    hash_password, verify_password,
    make_session_jwt, set_session_cookie, clear_session_cookie, get_session_token,
    get_current_user, get_current_user_optional, require_admin, require_roles, require_perm,
    check_lockout, record_failed_login, clear_login_attempts,
    fetch_emergent_profile, create_db_session, gen_reset_token,
    ADMIN_ROLES, ALL_PERMISSIONS,
)
from sl_locations import SL_DISTRICT_CITIES, all_districts, cities_for

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

app = FastAPI(title="ERP Storefront")
api = APIRouter(prefix="/api")


# ========== STARTUP ==========
async def _ensure_rls(conn):
    from sqlalchemy import text
    rows = (await conn.execute(text(
        "SELECT tablename FROM pg_tables WHERE schemaname='public' AND rowsecurity=false"
    ))).fetchall()
    for (t,) in rows:
        await conn.execute(text(f'ALTER TABLE public."{t}" ENABLE ROW LEVEL SECURITY;'))
        await conn.execute(text(f'ALTER TABLE public."{t}" FORCE ROW LEVEL SECURITY;'))
        await conn.execute(text(
            f'CREATE POLICY service_role_all ON public."{t}" FOR ALL TO service_role USING (true) WITH CHECK (true);'
        ))
        await conn.execute(text(
            f'CREATE POLICY deny_public ON public."{t}" FOR ALL TO anon, authenticated USING (false) WITH CHECK (false);'
        ))


# Idempotent column additions for evolving schema (run on startup).
COLUMN_MIGRATIONS = [
    ('users', 'permissions', 'JSONB'),
    ('products', 'supplier_id', 'VARCHAR(64)'),
    ('products', 'cost_price', 'DOUBLE PRECISION'),
    ('coupons', 'scope', "VARCHAR(16) DEFAULT 'all'"),
    ('coupons', 'scope_product_ids', 'JSONB'),
    ('coupons', 'scope_category_ids', 'JSONB'),
    ('expenses', 'store_id', 'VARCHAR(64)'),
    ('expenses', 'method', "VARCHAR(16) DEFAULT 'cash'"),
    ('expenses', 'cash_account_id', 'VARCHAR(64)'),
    ('orders', 'cash_tendered', 'DOUBLE PRECISION'),
    ('orders', 'cash_change', 'DOUBLE PRECISION'),
    ('orders', 'card_last4', 'VARCHAR(8)'),
    ('orders', 'cash_account_id', 'VARCHAR(64)'),
    ('company_settings', 'meta_title', 'VARCHAR(255)'),
    ('company_settings', 'meta_description', 'TEXT'),
    ('company_settings', 'meta_keywords', 'TEXT'),
    ('company_settings', 'og_image_id', 'VARCHAR(64)'),
    ('company_settings', 'google_analytics_id', 'VARCHAR(64)'),
    ('company_settings', 'google_site_verification', 'VARCHAR(128)'),
    ('company_settings', 'facebook_pixel_id', 'VARCHAR(64)'),
    ('company_settings', 'instagram_url', 'VARCHAR(255)'),
    ('company_settings', 'facebook_url', 'VARCHAR(255)'),
    ('company_settings', 'tiktok_url', 'VARCHAR(255)'),
    ('company_settings', 'twitter_url', 'VARCHAR(255)'),
    ('company_settings', 'youtube_url', 'VARCHAR(255)'),
]


async def _migrate_columns(conn):
    from sqlalchemy import text
    for table, col, ddl in COLUMN_MIGRATIONS:
        try:
            await conn.execute(text(f'ALTER TABLE public."{table}" ADD COLUMN IF NOT EXISTS "{col}" {ddl}'))
        except Exception as e:
            logger.warning(f"Migration skipped {table}.{col}: {e}")


DEFAULT_THEME = {
    "primary_color": "#FF3B30",
    "primary_color_hover": "#D92D23",
    "background_color": "#09090B",
    "text_color": "#FFFFFF",
    "text_muted_color": "#A1A1AA",
    "font_eyebrow": "'Space Grotesk', sans-serif",
    "font_heading": "'Archivo Black', sans-serif",
    "font_body": "'Inter', sans-serif",
    "heading_scale": 1.0,
    "line_height": 1.5,
    "marquee_phrases": ["NEW DROP", "EST. 2026"],
    "marquee_separator": "//",
}

DEFAULT_HEADER_CONFIG = {
    "menu": [
        {"label": "Home", "url": "/"},
        {"label": "Shop", "url": "/shop"},
        {"label": "Account", "url": "/account"},
    ],
    "show_search": True,
    "show_cart": True,
    "show_login": True,
    "style": "minimal",  # minimal, bold, classic
    "sticky": True,
}

DEFAULT_FOOTER_CONFIG = {
    "tagline": "",
    "support_email": "",
    "support_phone": "",
    "columns": [
        {"title": "Shop", "links": [{"label": "All Products", "url": "/shop"}, {"label": "Featured", "url": "/shop?featured=1"}]},
        {"title": "Support", "links": [{"label": "Shipping Policy", "url": "/page/shipping-policy"}, {"label": "Returns & Refunds", "url": "/page/returns-policy"}]},
        {"title": "Account", "links": [{"label": "Sign In", "url": "/login"}, {"label": "My Orders", "url": "/account"}]},
    ],
    "social_links": [],
    "copyright": "All rights reserved.",
    "show_marquee": True,
}

DEFAULT_HOME_SECTIONS = [
    {"section_type": "hero", "config": {
        "badge_text": "NEW SEASON IN STOCK",
        "headline_line1": "Welcome to your store.",
        "headline_line2": "Edit me anytime.",
        "headline_line2_accent": True,
        "headline_size": "lg",
        "subheading": "Tell your story here. Customize from Page Builder in your admin panel.",
        "cta_primary_label": "Shop The Collection",
        "cta_primary_link": "/shop",
        "cta_secondary_label": "View Our Story",
        "cta_secondary_link": "/page/about",
        "image_url": "",
        "image_id": None,
        "image_position": "right",
        "overlay_opacity": 60,
        "height": "tall",
    }},
    {"section_type": "featured", "config": {
        "eyebrow": "Featured", "heading": "Best Sellers", "max_items": 8,
        "category_slug": None, "show_view_all_button": True,
        "view_all_label": "Shop All", "view_all_link": "/shop",
    }},
    {"section_type": "brand", "config": {
        "eyebrow": "About Us", "headline": "Built right.\nWorn easy.\nMade to last.",
        "paragraph": "Tell customers what makes your brand different.",
        "stats": [{"value": "100%", "label": "Quality"}],
        "image_url": "", "image_id": None, "image_side": "right", "tagline": "",
    }},
]

DEFAULT_PRODUCT_PAGE_SECTIONS = [
    {"section_type": "featured", "config": {
        "eyebrow": "You May Also Like", "heading": "Same Category", "max_items": 4,
        "category_slug": "_same_category", "show_view_all_button": False,
    }},
]

DEFAULT_POLICY_SHIPPING = [
    {"section_type": "custom", "config": {
        "block_type": "heading_text", "eyebrow": "POLICIES",
        "heading": "Shipping Policy",
        "text": "Edit this page from Admin → Page Builder → Shipping Policy. Customize text, add sections, link products to it.",
        "alignment": "left", "max_width": "narrow", "padding": "lg",
    }},
]

DEFAULT_POLICY_RETURNS = [
    {"section_type": "custom", "config": {
        "block_type": "heading_text", "eyebrow": "POLICIES",
        "heading": "Returns & Refunds",
        "text": "Edit this page from Admin → Page Builder → Returns & Refunds.",
        "alignment": "left", "max_width": "narrow", "padding": "lg",
    }},
]

DEFAULT_PAYMENT_METHODS = [
    {"code": "cod", "label": "Cash on Delivery", "scope": "online", "active": True, "sort_order": 0,
     "description": "Pay with cash when your order is delivered."},
    {"code": "payhere", "label": "PayHere (Card / Bank)", "scope": "online", "active": False, "sort_order": 10,
     "description": "Secure online payment via PayHere gateway.",
     "config": {"merchant_id": "", "merchant_secret": "", "sandbox": True}},
    {"code": "koko", "label": "KOKO — Pay in 3 (Online)", "scope": "online", "active": False, "sort_order": 20,
     "description": "Buy now, pay in 3 interest-free instalments via KOKO.",
     "config": {"merchant_id": "", "api_key": "", "secret": "", "sandbox": True,
                 "_setup_help": "Add your KOKO merchant credentials. Until live SDK is wired, orders mark as paid."}},
    {"code": "mintpay", "label": "Mintpay — Buy Now Pay Later", "scope": "online", "active": False, "sort_order": 30,
     "description": "Split your order into instalments with Mintpay.",
     "config": {"merchant_id": "", "api_key": "", "secret": "", "sandbox": True,
                 "_setup_help": "Add your Mintpay merchant credentials. Until live SDK is wired, orders mark as paid."}},
    {"code": "cash", "label": "Cash", "scope": "pos", "active": True, "sort_order": 0},
    {"code": "card_pos", "label": "Card (POS Terminal)", "scope": "pos", "active": True, "sort_order": 10},
    {"code": "koko_pos", "label": "KOKO — Pay in 3 (POS)", "scope": "pos", "active": False, "sort_order": 20,
     "description": "Customer pays via KOKO QR on the POS counter.",
     "config": {"merchant_id": "", "api_key": "", "secret": "", "sandbox": True}},
    {"code": "mintpay_pos", "label": "Mintpay (POS)", "scope": "pos", "active": False, "sort_order": 30,
     "config": {"merchant_id": "", "api_key": "", "secret": "", "sandbox": True}},
]

DEFAULT_SYSTEM_PAGES = [
    {"slug": "home", "title": "Home", "is_system": True, "show_in_header_menu": False, "sort_order": 0},
    {"slug": "_header", "title": "Site Header", "is_system": True, "show_in_header_menu": False, "sort_order": -100},
    {"slug": "_footer", "title": "Site Footer", "is_system": True, "show_in_header_menu": False, "sort_order": -90},
    {"slug": "_product_page", "title": "Product Page Layout", "is_system": True, "show_in_header_menu": False, "sort_order": -80},
    {"slug": "shipping-policy", "title": "Shipping Policy", "is_system": True, "show_in_header_menu": False, "sort_order": 100},
    {"slug": "returns-policy", "title": "Returns & Refunds", "is_system": True, "show_in_header_menu": False, "sort_order": 110},
]


@app.on_event("startup")
async def startup():
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
        await _migrate_columns(conn)
        await _ensure_rls(conn)
    from database import AsyncSessionLocal
    async with AsyncSessionLocal() as db:
        # Default store
        store = (await db.execute(select(M.Store))).scalars().first()
        if not store:
            db.add(M.Store(name="Online Store", is_online=True, active=True, address="Online"))
            await db.commit()
        # Theme
        theme = (await db.execute(select(M.ThemeSetting).where(M.ThemeSetting.id == "default"))).scalar_one_or_none()
        if not theme:
            db.add(M.ThemeSetting(id="default", config=DEFAULT_THEME))
            await db.commit()
        # Company settings
        cs = (await db.execute(select(M.CompanySettings).where(M.CompanySettings.id == "default"))).scalar_one_or_none()
        if not cs:
            db.add(M.CompanySettings(id="default", company_name="My Brand", currency="LKR"))
            await db.commit()
        # System pages
        existing_pages = {p.slug for p in (await db.execute(select(M.CustomPage))).scalars().all()}
        for sp in DEFAULT_SYSTEM_PAGES:
            if sp["slug"] not in existing_pages:
                db.add(M.CustomPage(**sp))
        await db.commit()
        # Default sections
        async def _seed_page(slug, sections):
            existing = (await db.execute(select(M.PageSection).where(M.PageSection.page == slug))).scalars().first()
            if not existing:
                for i, sec in enumerate(sections):
                    db.add(M.PageSection(page=slug, section_type=sec["section_type"],
                                         sort_order=i * 10, visible=True, config=sec["config"]))
        await _seed_page("home", DEFAULT_HOME_SECTIONS)
        await _seed_page("_product_page", DEFAULT_PRODUCT_PAGE_SECTIONS)
        await _seed_page("shipping-policy", DEFAULT_POLICY_SHIPPING)
        await _seed_page("returns-policy", DEFAULT_POLICY_RETURNS)
        # Header/footer config: stored as a single 'config_block' section
        for slug, default_cfg in [("_header", DEFAULT_HEADER_CONFIG), ("_footer", DEFAULT_FOOTER_CONFIG)]:
            existing = (await db.execute(select(M.PageSection).where(M.PageSection.page == slug))).scalars().first()
            if not existing:
                db.add(M.PageSection(page=slug, section_type="config_block", sort_order=0, visible=True, config=default_cfg))
        # Payment methods
        if not (await db.execute(select(M.PaymentMethod))).scalars().first():
            for pm in DEFAULT_PAYMENT_METHODS:
                db.add(M.PaymentMethod(**pm))
        await db.commit()


# ========== UTILS ==========
def slugify(text: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "-", text.strip().lower()).strip("-")
    return s or uuid.uuid4().hex[:8]


def new_order_number() -> str:
    return "ORD-" + datetime.now(timezone.utc).strftime("%Y%m%d") + "-" + "".join(random.choices(string.ascii_uppercase + string.digits, k=5))


def normalize_phone_lk(raw):
    """Sri Lanka phone normalisation -> E.164. Accepts 0777..., 777..., 94777..., +94777... ."""
    if not raw:
        return raw
    digits = re.sub(r"[^\d+]", "", str(raw))
    if digits.startswith("+94"): return digits
    if digits.startswith("94"): return f"+{digits}"
    if digits.startswith("0"): return f"+94{digits[1:]}"
    if re.fullmatch(r"[1-9]\d{8}", digits): return f"+94{digits}"
    return digits if digits.startswith("+") else digits


async def _select_active_discounts(db: AsyncSession):
    now = datetime.now(timezone.utc)
    rows = (await db.execute(select(M.Discount).where(M.Discount.active == True))).scalars().all()
    out = []
    for d in rows:
        if d.starts_at:
            sa = d.starts_at if d.starts_at.tzinfo else d.starts_at.replace(tzinfo=timezone.utc)
            if sa > now: continue
        if d.ends_at:
            ea = d.ends_at if d.ends_at.tzinfo else d.ends_at.replace(tzinfo=timezone.utc)
            if ea < now: continue
        out.append(d)
    return out


def _best_discount_for(product, unit_price, discounts):
    best_save = 0.0
    best_d = None
    for d in discounts:
        if d.scope == "sitewide":
            applies = True
        elif d.scope == "products" and product.id in (d.scope_product_ids or []):
            applies = True
        elif d.scope == "categories" and product.category_id and product.category_id in (d.scope_category_ids or []):
            applies = True
        else:
            applies = False
        if not applies: continue
        save = unit_price * (d.value / 100.0) if d.type == "percent" else min(unit_price, d.value)
        if save > best_save:
            best_save = save; best_d = d
    return best_save, best_d


def _client_ip(request: Request) -> str:
    fwd = request.headers.get("x-forwarded-for", "")
    if fwd:
        return fwd.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def _public_user(u: M.User) -> dict:
    return {"user_id": u.user_id, "email": u.email, "name": u.name, "picture": u.picture,
            "role": u.role, "phone": u.phone, "permissions": u.permissions or {}}


# ========== SETUP WIZARD ==========
class SetupInit(BaseModel):
    company_name: str = Field(..., min_length=1, max_length=255)
    tagline: Optional[str] = None
    company_email: Optional[str] = None
    company_phone: Optional[str] = None
    company_address: Optional[str] = None
    currency: str = "LKR"
    admin_email: EmailStr
    admin_name: str = Field(..., min_length=1)
    admin_password: str = Field(..., min_length=8)


@api.get("/setup/status")
async def setup_status(db: AsyncSession = Depends(get_db)):
    cs = (await db.execute(select(M.CompanySettings).where(M.CompanySettings.id == "default"))).scalar_one_or_none()
    return {"setup_complete": bool(cs and cs.setup_complete),
            "company_name": cs.company_name if cs else None}


@api.post("/setup/init")
async def setup_init(payload: SetupInit, response: Response, db: AsyncSession = Depends(get_db)):
    cs = (await db.execute(select(M.CompanySettings).where(M.CompanySettings.id == "default"))).scalar_one_or_none()
    if cs and cs.setup_complete:
        raise HTTPException(409, "Setup already completed")
    # Create the super_admin
    email = payload.admin_email.lower()
    existing = (await db.execute(select(M.User).where(M.User.email == email))).scalar_one_or_none()
    if existing:
        existing.password_hash = hash_password(payload.admin_password)
        existing.name = payload.admin_name
        existing.role = "super_admin"
        existing.active = True
        existing.auth_provider = "password"
        admin = existing
    else:
        admin = M.User(
            user_id=f"user_{uuid.uuid4().hex[:12]}",
            email=email, name=payload.admin_name, role="super_admin",
            password_hash=hash_password(payload.admin_password),
            auth_provider="password", active=True,
        )
        db.add(admin)
    if not cs:
        cs = M.CompanySettings(id="default")
        db.add(cs)
    cs.company_name = payload.company_name
    cs.tagline = payload.tagline
    cs.email = payload.company_email
    cs.phone = payload.company_phone
    cs.address = payload.company_address
    cs.currency = payload.currency or "LKR"
    cs.setup_complete = True
    cs.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(admin)
    token = make_session_jwt(admin.user_id)
    set_session_cookie(response, token)
    return {"ok": True, "user": _public_user(admin), "token": token}


# ========== AUTH ==========
class LoginIn(BaseModel):
    email: EmailStr
    password: str


class RegisterIn(BaseModel):
    email: EmailStr
    password: str = Field(..., min_length=8)
    name: str
    phone: Optional[str] = None


class ChangePasswordIn(BaseModel):
    current_password: str
    new_password: str = Field(..., min_length=8)


@api.post("/auth/login")
async def login(payload: LoginIn, request: Request, response: Response, db: AsyncSession = Depends(get_db)):
    email = payload.email.lower()
    ident = f"{_client_ip(request)}:{email}"
    await check_lockout(db, ident)
    user = (await db.execute(select(M.User).where(M.User.email == email))).scalar_one_or_none()
    if not user or not user.password_hash or not verify_password(payload.password, user.password_hash):
        await record_failed_login(db, ident)
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if not user.active:
        raise HTTPException(status_code=403, detail="Account disabled")
    await clear_login_attempts(db, ident)
    token = make_session_jwt(user.user_id)
    set_session_cookie(response, token)
    return {"user": _public_user(user), "token": token}


@api.post("/auth/register")
async def register(payload: RegisterIn, response: Response, db: AsyncSession = Depends(get_db)):
    email = payload.email.lower()
    existing = (await db.execute(select(M.User).where(M.User.email == email))).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="Email already registered")
    user = M.User(
        user_id=f"user_{uuid.uuid4().hex[:12]}",
        email=email, name=payload.name, phone=payload.phone,
        password_hash=hash_password(payload.password),
        auth_provider="password", role="customer", active=True,
    )
    db.add(user)
    await db.flush()
    db.add(M.Customer(user_id=user.user_id, name=user.name, email=user.email, phone=user.phone))
    await db.commit()
    await db.refresh(user)
    token = make_session_jwt(user.user_id)
    set_session_cookie(response, token)
    return {"user": _public_user(user), "token": token}


@api.get("/auth/me")
async def auth_me(user: M.User = Depends(get_current_user)):
    return _public_user(user)


@api.post("/auth/logout")
async def logout(request: Request, response: Response, db: AsyncSession = Depends(get_db)):
    token = await get_session_token(request)
    if token:
        # If this is a DB-backed session (Google OAuth), remove it
        sess = (await db.execute(select(M.UserSession).where(M.UserSession.session_token == token))).scalar_one_or_none()
        if sess:
            await db.delete(sess)
            await db.commit()
    clear_session_cookie(response)
    return {"ok": True}


@api.post("/auth/change-password")
async def change_password(payload: ChangePasswordIn, user: M.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if not user.password_hash or not verify_password(payload.current_password, user.password_hash):
        raise HTTPException(400, "Current password is incorrect")
    user.password_hash = hash_password(payload.new_password)
    user.auth_provider = "password"
    await db.commit()
    return {"ok": True}


# ---- Google OAuth (customer-only) ----
class SessionRequest(BaseModel):
    session_id: str


@api.post("/auth/session")
async def auth_session(payload: SessionRequest, response: Response, db: AsyncSession = Depends(get_db)):
    profile = await fetch_emergent_profile(payload.session_id)
    email = (profile.get("email") or "").lower()
    if not email:
        raise HTTPException(status_code=400, detail="Missing email")
    user = (await db.execute(select(M.User).where(M.User.email == email))).scalar_one_or_none()
    if user:
        # Existing user - update profile picture/name only
        if profile.get("name"):
            user.name = profile["name"]
        if profile.get("picture"):
            user.picture = profile["picture"]
    else:
        # Brand new user via Google = always customer (admins must use password)
        user = M.User(
            user_id=f"user_{uuid.uuid4().hex[:12]}",
            email=email, name=profile.get("name", email.split("@")[0]),
            picture=profile.get("picture"), role="customer",
            auth_provider="google", active=True,
        )
        db.add(user)
        await db.flush()
        db.add(M.Customer(user_id=user.user_id, name=user.name, email=user.email))
    await db.commit()
    await db.refresh(user)
    # Use DB-backed session for Google flow (matches their session_token)
    sess_token = profile.get("session_token") or uuid.uuid4().hex
    await create_db_session(db, user, sess_token)
    set_session_cookie(response, sess_token)
    return {"user": _public_user(user)}


# ========== COMPANY / SETTINGS ==========
def _company_dict(cs: M.CompanySettings):
    return {
        "company_name": cs.company_name, "tagline": cs.tagline,
        "email": cs.email, "phone": cs.phone, "address": cs.address,
        "currency": cs.currency,
        "logo_light_id": cs.logo_light_id, "logo_dark_id": cs.logo_dark_id,
        "favicon_id": cs.favicon_id, "setup_complete": cs.setup_complete,
        "meta_title": cs.meta_title, "meta_description": cs.meta_description,
        "meta_keywords": cs.meta_keywords, "og_image_id": cs.og_image_id,
        "google_analytics_id": cs.google_analytics_id,
        "google_site_verification": cs.google_site_verification,
        "facebook_pixel_id": cs.facebook_pixel_id,
        "instagram_url": cs.instagram_url, "facebook_url": cs.facebook_url,
        "tiktok_url": cs.tiktok_url, "twitter_url": cs.twitter_url, "youtube_url": cs.youtube_url,
    }


class CompanyUpdate(BaseModel):
    company_name: Optional[str] = None
    tagline: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    currency: Optional[str] = None
    logo_light_id: Optional[str] = None
    logo_dark_id: Optional[str] = None
    favicon_id: Optional[str] = None
    meta_title: Optional[str] = None
    meta_description: Optional[str] = None
    meta_keywords: Optional[str] = None
    og_image_id: Optional[str] = None
    google_analytics_id: Optional[str] = None
    google_site_verification: Optional[str] = None
    facebook_pixel_id: Optional[str] = None
    instagram_url: Optional[str] = None
    facebook_url: Optional[str] = None
    tiktok_url: Optional[str] = None
    twitter_url: Optional[str] = None
    youtube_url: Optional[str] = None


@api.get("/company")
async def get_company(db: AsyncSession = Depends(get_db)):
    cs = (await db.execute(select(M.CompanySettings).where(M.CompanySettings.id == "default"))).scalar_one_or_none()
    if not cs:
        return {"company_name": "My Brand", "currency": "LKR", "setup_complete": False}
    return _company_dict(cs)


@api.put("/admin/company")
async def update_company(payload: CompanyUpdate, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    cs = (await db.execute(select(M.CompanySettings).where(M.CompanySettings.id == "default"))).scalar_one_or_none()
    if not cs:
        cs = M.CompanySettings(id="default")
        db.add(cs)
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(cs, k, v)
    cs.updated_at = datetime.now(timezone.utc)
    await db.commit()
    return _company_dict(cs)


# ========== INTEGRATION SETTINGS (SMTP/SendGrid/Brevo/Twilio/Notify.lk) ==========
class IntegrationIn(BaseModel):
    kind: str  # email, sms
    provider: str
    label: Optional[str] = None
    config: dict = {}
    active: bool = False
    is_default: bool = False


def _integration_dict(i: M.IntegrationSetting):
    return {"id": i.id, "kind": i.kind, "provider": i.provider, "label": i.label,
            "config": i.config or {}, "active": i.active, "is_default": i.is_default,
            "created_at": i.created_at.isoformat()}


@api.get("/admin/integrations")
async def list_integrations(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    rows = (await db.execute(select(M.IntegrationSetting).order_by(M.IntegrationSetting.kind, M.IntegrationSetting.created_at))).scalars().all()
    return [_integration_dict(i) for i in rows]


@api.post("/admin/integrations")
async def create_integration(payload: IntegrationIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_roles("super_admin"))):
    if payload.is_default:
        existing = (await db.execute(select(M.IntegrationSetting).where(M.IntegrationSetting.kind == payload.kind))).scalars().all()
        for e in existing:
            e.is_default = False
    i = M.IntegrationSetting(**payload.model_dump())
    db.add(i)
    await db.commit()
    await db.refresh(i)
    return _integration_dict(i)


@api.put("/admin/integrations/{iid}")
async def update_integration(iid: str, payload: IntegrationIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_roles("super_admin"))):
    i = (await db.execute(select(M.IntegrationSetting).where(M.IntegrationSetting.id == iid))).scalar_one_or_none()
    if not i:
        raise HTTPException(404, "Not found")
    if payload.is_default:
        for e in (await db.execute(select(M.IntegrationSetting).where(and_(M.IntegrationSetting.kind == payload.kind, M.IntegrationSetting.id != iid)))).scalars().all():
            e.is_default = False
    for k, v in payload.model_dump().items():
        setattr(i, k, v)
    await db.commit()
    return _integration_dict(i)


@api.delete("/admin/integrations/{iid}")
async def delete_integration(iid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_roles("super_admin"))):
    i = (await db.execute(select(M.IntegrationSetting).where(M.IntegrationSetting.id == iid))).scalar_one_or_none()
    if not i:
        raise HTTPException(404, "Not found")
    await db.delete(i)
    await db.commit()
    return {"ok": True}


# ========== CATEGORIES ==========
class CategoryIn(BaseModel):
    name: str
    description: Optional[str] = None
    parent_id: Optional[str] = None
    sort_order: int = 0


def cat_to_dict(c):
    return {"id": c.id, "name": c.name, "slug": c.slug, "description": c.description,
            "parent_id": c.parent_id, "sort_order": c.sort_order}


def _build_category_tree(rows):
    by_id = {c["id"]: {**c, "children": []} for c in rows}
    roots = []
    for c in by_id.values():
        if c["parent_id"] and c["parent_id"] in by_id:
            by_id[c["parent_id"]]["children"].append(c)
        else:
            roots.append(c)
    return roots


def _descendant_ids(rows, parent_id):
    """Recursively collect descendant category ids (including self)."""
    out = {parent_id}
    children = [r for r in rows if r["parent_id"] == parent_id]
    for c in children:
        out.update(_descendant_ids(rows, c["id"]))
    return out


@api.get("/categories")
async def list_categories(db: AsyncSession = Depends(get_db), q: Optional[str] = None, tree: bool = False):
    query = select(M.Category).order_by(M.Category.sort_order, M.Category.name)
    if q:
        query = query.where(M.Category.name.ilike(f"%{q}%"))
    rows = (await db.execute(query)).scalars().all()
    flat = [cat_to_dict(c) for c in rows]
    if tree:
        return _build_category_tree(flat)
    return flat


@api.post("/admin/categories")
async def create_category(payload: CategoryIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    c = M.Category(name=payload.name, slug=slugify(payload.name), description=payload.description,
                    parent_id=payload.parent_id or None, sort_order=payload.sort_order)
    db.add(c)
    await db.commit()
    await db.refresh(c)
    return cat_to_dict(c)


@api.put("/admin/categories/{cid}")
async def update_category(cid: str, payload: CategoryIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    c = (await db.execute(select(M.Category).where(M.Category.id == cid))).scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Not found")
    if payload.parent_id == cid:
        raise HTTPException(400, "Cannot be own parent")
    c.name = payload.name
    c.slug = slugify(payload.name)
    c.description = payload.description
    c.parent_id = payload.parent_id or None
    c.sort_order = payload.sort_order
    await db.commit()
    return cat_to_dict(c)


@api.delete("/admin/categories/{cid}")
async def delete_category(cid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    c = (await db.execute(select(M.Category).where(M.Category.id == cid))).scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Not found")
    await db.delete(c)
    await db.commit()
    return {"ok": True}


# ========== PRODUCTS ==========
class VariantIn(BaseModel):
    id: Optional[str] = None
    size: Optional[str] = None
    color: Optional[str] = None
    color_hex: Optional[str] = None
    price_override: Optional[float] = None
    sku: Optional[str] = None
    stock: int = 0


class ProductIn(BaseModel):
    name: str
    description: Optional[str] = None
    category_id: Optional[str] = None
    supplier_id: Optional[str] = None
    base_price: float
    compare_price: Optional[float] = None
    cost_price: Optional[float] = None
    sku: Optional[str] = None
    status: str = "active"
    featured: bool = False
    shipping_note: Optional[str] = None
    returns_note: Optional[str] = None
    variants: List[VariantIn] = []


async def product_to_dict(p: M.Product, db: AsyncSession, include_details: bool = True):
    data = {
        "id": p.id, "name": p.name, "slug": p.slug, "description": p.description,
        "category_id": p.category_id, "supplier_id": p.supplier_id,
        "base_price": p.base_price, "compare_price": p.compare_price, "cost_price": p.cost_price,
        "sku": p.sku, "status": p.status, "featured": p.featured,
        "shipping_note": p.shipping_note, "returns_note": p.returns_note,
        "created_at": p.created_at.isoformat(),
    }
    imgs_q = await db.execute(
        select(M.ProductImage.id, M.ProductImage.mime_type, M.ProductImage.is_primary, M.ProductImage.sort_order, M.ProductImage.color)
        .where(M.ProductImage.product_id == p.id).order_by(M.ProductImage.sort_order)
    )
    imgs = imgs_q.all()
    if include_details:
        data["images"] = [{"id": i.id, "url": f"/api/images/{i.id}", "mime_type": i.mime_type,
                          "is_primary": i.is_primary, "color": i.color} for i in imgs]
    else:
        primary = [i for i in imgs if i.is_primary] or (imgs[:1] if imgs else [])
        data["images"] = [{"id": i.id, "url": f"/api/images/{i.id}", "mime_type": i.mime_type,
                          "is_primary": True, "color": i.color} for i in primary]
    if include_details:
        variants = (await db.execute(select(M.Variant).where(M.Variant.product_id == p.id))).scalars().all()
        inv_map = {}
        if variants:
            for iv in (await db.execute(select(M.Inventory).where(M.Inventory.variant_id.in_([v.id for v in variants])))).scalars().all():
                inv_map[iv.variant_id] = inv_map.get(iv.variant_id, 0) + iv.quantity
        data["variants"] = [{"id": v.id, "size": v.size, "color": v.color, "color_hex": v.color_hex,
                            "price_override": v.price_override, "sku": v.sku, "stock": inv_map.get(v.id, 0)}
                           for v in variants]
        cat = (await db.execute(select(M.Category).where(M.Category.id == p.category_id))).scalar_one_or_none() if p.category_id else None
        data["category"] = {"id": cat.id, "name": cat.name, "slug": cat.slug} if cat else None
    return data


async def products_to_list(rows, db: AsyncSession):
    if not rows:
        return []
    pids = [p.id for p in rows]
    imgs_q = await db.execute(
        select(M.ProductImage.id, M.ProductImage.product_id, M.ProductImage.mime_type, M.ProductImage.is_primary, M.ProductImage.sort_order, M.ProductImage.color)
        .where(M.ProductImage.product_id.in_(pids)).order_by(M.ProductImage.sort_order)
    )
    imgs_by_pid = {}
    for i in imgs_q.all():
        imgs_by_pid.setdefault(i.product_id, []).append(i)
    cat_ids = list({p.category_id for p in rows if p.category_id})
    cats = {}
    if cat_ids:
        for c in (await db.execute(select(M.Category).where(M.Category.id.in_(cat_ids)))).scalars().all():
            cats[c.id] = {"id": c.id, "name": c.name, "slug": c.slug}
    out = []
    for p in rows:
        pimgs = imgs_by_pid.get(p.id, [])
        primary = [i for i in pimgs if i.is_primary] or (pimgs[:1] if pimgs else [])
        out.append({
            "id": p.id, "name": p.name, "slug": p.slug, "base_price": p.base_price,
            "compare_price": p.compare_price, "featured": p.featured, "status": p.status,
            "category": cats.get(p.category_id),
            "images": [{"id": i.id, "url": f"/api/images/{i.id}", "mime_type": i.mime_type, "is_primary": True, "color": i.color} for i in primary],
        })
    return out


@api.get("/products")
async def list_products(
    db: AsyncSession = Depends(get_db),
    category: Optional[str] = None, featured: Optional[bool] = None,
    search: Optional[str] = None, limit: int = 50, exclude_id: Optional[str] = None,
):
    q = select(M.Product).where(M.Product.status == "active")
    if category:
        # Resolve category slug + all descendant categories
        cat = (await db.execute(select(M.Category).where(M.Category.slug == category))).scalar_one_or_none()
        if cat:
            all_cats = (await db.execute(select(M.Category))).scalars().all()
            flat = [{"id": c.id, "parent_id": c.parent_id} for c in all_cats]
            ids = _descendant_ids(flat, cat.id)
            q = q.where(M.Product.category_id.in_(list(ids)))
    if featured:
        q = q.where(M.Product.featured == True)
    if search:
        q = q.where(M.Product.name.ilike(f"%{search}%"))
    if exclude_id:
        q = q.where(M.Product.id != exclude_id)
    q = q.order_by(desc(M.Product.created_at)).limit(limit)
    rows = (await db.execute(q)).scalars().all()
    return await products_to_list(rows, db)


@api.get("/products/{slug}")
async def get_product(slug: str, db: AsyncSession = Depends(get_db)):
    p = (await db.execute(select(M.Product).where(M.Product.slug == slug))).scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Not found")
    return await product_to_dict(p, db, include_details=True)


@api.get("/admin/products")
async def admin_list_products(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin),
                               q: Optional[str] = None, page: int = 1, page_size: int = 50,
                               store_id: Optional[str] = None, in_stock: bool = False):
    base = select(M.Product)
    if q:
        base = base.where(or_(M.Product.name.ilike(f"%{q}%"), M.Product.sku.ilike(f"%{q}%")))
    total = (await db.execute(select(func.count()).select_from(base.subquery()))).scalar_one()
    page_size = min(max(1, page_size), 100); page = max(1, page)
    rows = (await db.execute(base.order_by(desc(M.Product.created_at)).offset((page-1)*page_size).limit(page_size))).scalars().all()
    items = [await product_to_dict(p, db, include_details=True) for p in rows]
    # Optional store-stock filter (used by POS): only keep variants with positive stock at the chosen store
    if store_id and in_stock:
        invs = (await db.execute(select(M.Inventory).where(M.Inventory.store_id == store_id))).scalars().all()
        stock_map = {iv.variant_id: iv.quantity for iv in invs}
        filtered = []
        for p in items:
            keep_variants = []
            for v in p.get("variants", []):
                qty = stock_map.get(v["id"], 0)
                v["stock"] = qty
                if qty > 0:
                    keep_variants.append(v)
            if keep_variants:
                p["variants"] = keep_variants
                filtered.append(p)
        items = filtered
        total = len(items)
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@api.get("/images/{img_id}")
async def stream_image(img_id: str, db: AsyncSession = Depends(get_db)):
    row = (await db.execute(
        select(M.ProductImage.data_base64, M.ProductImage.mime_type).where(M.ProductImage.id == img_id)
    )).first()
    if not row:
        raise HTTPException(404, "Not found")
    try:
        raw = _b64.b64decode(row.data_base64)
    except Exception:
        raise HTTPException(500, "Corrupt image")
    return FastResponse(content=raw, media_type=row.mime_type or "image/png",
                        headers={"Cache-Control": "public, max-age=31536000, immutable"})


async def _ensure_default_store(db: AsyncSession) -> M.Store:
    store = (await db.execute(select(M.Store).where(M.Store.is_online == True))).scalars().first()
    if not store:
        store = (await db.execute(select(M.Store))).scalars().first()
    if not store:
        store = M.Store(name="Online Store", is_online=True)
        db.add(store)
        await db.commit()
        await db.refresh(store)
    return store


@api.post("/admin/products")
async def create_product(payload: ProductIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    store = await _ensure_default_store(db)
    slug = slugify(payload.name) + "-" + uuid.uuid4().hex[:4]
    p = M.Product(
        name=payload.name, slug=slug, description=payload.description,
        category_id=payload.category_id, supplier_id=payload.supplier_id,
        base_price=payload.base_price, compare_price=payload.compare_price,
        cost_price=payload.cost_price, sku=payload.sku,
        status=payload.status, featured=payload.featured,
        shipping_note=payload.shipping_note, returns_note=payload.returns_note,
    )
    db.add(p)
    await db.commit()
    await db.refresh(p)
    for v in payload.variants:
        variant = M.Variant(product_id=p.id, size=v.size, color=v.color, color_hex=v.color_hex,
                            price_override=v.price_override, sku=v.sku)
        db.add(variant)
        await db.flush()
        db.add(M.Inventory(variant_id=variant.id, store_id=store.id, quantity=v.stock))
    await db.commit()
    return await product_to_dict(p, db, include_details=True)


@api.put("/admin/products/{pid}")
async def update_product(pid: str, payload: ProductIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    p = (await db.execute(select(M.Product).where(M.Product.id == pid))).scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Not found")
    store = await _ensure_default_store(db)
    p.name = payload.name; p.description = payload.description
    p.category_id = payload.category_id; p.supplier_id = payload.supplier_id
    p.base_price = payload.base_price
    p.compare_price = payload.compare_price; p.cost_price = payload.cost_price
    p.sku = payload.sku
    p.status = payload.status; p.featured = payload.featured
    p.shipping_note = payload.shipping_note; p.returns_note = payload.returns_note
    p.updated_at = datetime.now(timezone.utc)
    existing_variants = (await db.execute(select(M.Variant).where(M.Variant.product_id == p.id))).scalars().all()
    existing_map = {v.id: v for v in existing_variants}
    incoming_ids = set()
    for v in payload.variants:
        if v.id and v.id in existing_map:
            ev = existing_map[v.id]
            ev.size = v.size; ev.color = v.color; ev.color_hex = v.color_hex
            ev.price_override = v.price_override; ev.sku = v.sku
            inv = (await db.execute(select(M.Inventory).where(and_(M.Inventory.variant_id == ev.id, M.Inventory.store_id == store.id)))).scalar_one_or_none()
            if inv:
                inv.quantity = v.stock
            else:
                db.add(M.Inventory(variant_id=ev.id, store_id=store.id, quantity=v.stock))
            incoming_ids.add(v.id)
        else:
            nv = M.Variant(product_id=p.id, size=v.size, color=v.color, color_hex=v.color_hex,
                          price_override=v.price_override, sku=v.sku)
            db.add(nv); await db.flush()
            db.add(M.Inventory(variant_id=nv.id, store_id=store.id, quantity=v.stock))
            incoming_ids.add(nv.id)
    for vid, v in existing_map.items():
        if vid not in incoming_ids:
            await db.delete(v)
    await db.commit()
    return await product_to_dict(p, db, include_details=True)


@api.delete("/admin/products/{pid}")
async def delete_product(pid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    p = (await db.execute(select(M.Product).where(M.Product.id == pid))).scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Not found")
    await db.delete(p)
    await db.commit()
    return {"ok": True}


class ImagePayload(BaseModel):
    data_base64: str
    mime_type: str = "image/png"
    is_primary: bool = False
    color: Optional[str] = None  # bind to a specific color variant


@api.post("/admin/products/{pid}/images")
async def add_product_image(pid: str, payload: ImagePayload, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    p = (await db.execute(select(M.Product).where(M.Product.id == pid))).scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Not found")
    if payload.is_primary:
        for e in (await db.execute(select(M.ProductImage).where(M.ProductImage.product_id == pid))).scalars().all():
            e.is_primary = False
    img = M.ProductImage(product_id=pid, data_base64=payload.data_base64, mime_type=payload.mime_type,
                         is_primary=payload.is_primary, color=payload.color)
    db.add(img)
    await db.commit()
    await db.refresh(img)
    return {"id": img.id, "is_primary": img.is_primary, "color": img.color}


class ImageMetaUpdate(BaseModel):
    is_primary: Optional[bool] = None
    color: Optional[str] = None
    sort_order: Optional[int] = None


@api.put("/admin/products/images/{img_id}")
async def update_image_meta(img_id: str, payload: ImageMetaUpdate, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    img = (await db.execute(select(M.ProductImage).where(M.ProductImage.id == img_id))).scalar_one_or_none()
    if not img:
        raise HTTPException(404, "Not found")
    if payload.is_primary:
        for e in (await db.execute(select(M.ProductImage).where(and_(M.ProductImage.product_id == img.product_id, M.ProductImage.id != img_id)))).scalars().all():
            e.is_primary = False
        img.is_primary = True
    elif payload.is_primary is False:
        img.is_primary = False
    if payload.color is not None:
        img.color = payload.color or None
    if payload.sort_order is not None:
        img.sort_order = payload.sort_order
    await db.commit()
    return {"id": img.id, "is_primary": img.is_primary, "color": img.color}


@api.delete("/admin/products/images/{img_id}")
async def delete_image(img_id: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    img = (await db.execute(select(M.ProductImage).where(M.ProductImage.id == img_id))).scalar_one_or_none()
    if not img:
        raise HTTPException(404, "Not found")
    await db.delete(img)
    await db.commit()
    return {"ok": True}


# ========== INVENTORY ==========
@api.get("/admin/inventory")
async def list_inventory(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin),
                         q: Optional[str] = None, store_id: Optional[str] = None,
                         page: int = 1, page_size: int = 50):
    rows = (await db.execute(select(M.Inventory))).scalars().all()
    out = []
    for iv in rows:
        if store_id and iv.store_id != store_id:
            continue
        v = (await db.execute(select(M.Variant).where(M.Variant.id == iv.variant_id))).scalar_one_or_none()
        p = (await db.execute(select(M.Product).where(M.Product.id == v.product_id))).scalar_one_or_none() if v else None
        s = (await db.execute(select(M.Store).where(M.Store.id == iv.store_id))).scalar_one_or_none()
        if q and p:
            q_low = q.lower()
            if q_low not in (p.name or "").lower() and q_low not in (v.sku or "").lower() and q_low not in (v.color or "").lower():
                continue
        out.append({
            "id": iv.id, "variant_id": iv.variant_id, "store_id": iv.store_id,
            "store_name": s.name if s else "", "quantity": iv.quantity,
            "low_stock_threshold": iv.low_stock_threshold,
            "product_name": p.name if p else "", "product_id": p.id if p else None,
            "variant_label": f"{v.size or ''} / {v.color or ''}" if v else "",
            "sku": v.sku if v else None, "low": iv.quantity <= iv.low_stock_threshold,
            "_created": v.created_at.isoformat() if v else "",
        })
    out.sort(key=lambda r: r.get("_created", ""), reverse=True)
    total = len(out)
    page_size = min(max(1, page_size), 100); page = max(1, page)
    start = (page - 1) * page_size
    items = out[start:start + page_size]
    return {"items": items, "total": total, "page": page, "page_size": page_size}


class StockMoveIn(BaseModel):
    variant_id: str
    store_id: Optional[str] = None
    type: str
    quantity: int
    reason: Optional[str] = None
    reference: Optional[str] = None


@api.post("/admin/stock-movements")
async def create_stock_movement(payload: StockMoveIn, db: AsyncSession = Depends(get_db), user: M.User = Depends(require_perm("move_stocks"))):
    store = await _ensure_default_store(db)
    sid = payload.store_id or store.id
    inv = (await db.execute(select(M.Inventory).where(and_(M.Inventory.variant_id == payload.variant_id, M.Inventory.store_id == sid)))).scalar_one_or_none()
    if not inv:
        inv = M.Inventory(variant_id=payload.variant_id, store_id=sid, quantity=0)
        db.add(inv); await db.flush()
    if payload.type == "in":
        inv.quantity += payload.quantity
    elif payload.type == "out":
        inv.quantity = max(0, inv.quantity - payload.quantity)
    elif payload.type == "adjust":
        inv.quantity = payload.quantity
    db.add(M.StockMovement(variant_id=payload.variant_id, store_id=sid, type=payload.type,
                            quantity=payload.quantity, reason=payload.reason,
                            reference=payload.reference, user_id=user.user_id))
    await db.commit()
    return {"ok": True, "new_quantity": inv.quantity}


class TransferIn(BaseModel):
    variant_id: str
    from_store_id: str
    to_store_id: str
    quantity: int
    reason: Optional[str] = None


@api.post("/admin/inventory/transfer")
async def transfer_stock(payload: TransferIn, db: AsyncSession = Depends(get_db), user: M.User = Depends(require_perm("move_stocks"))):
    if payload.from_store_id == payload.to_store_id:
        raise HTTPException(400, "Same source and destination")
    if payload.quantity <= 0:
        raise HTTPException(400, "Quantity must be > 0")
    src = (await db.execute(select(M.Inventory).where(and_(M.Inventory.variant_id == payload.variant_id, M.Inventory.store_id == payload.from_store_id)))).scalar_one_or_none()
    if not src or src.quantity < payload.quantity:
        raise HTTPException(400, "Insufficient stock at source")
    dst = (await db.execute(select(M.Inventory).where(and_(M.Inventory.variant_id == payload.variant_id, M.Inventory.store_id == payload.to_store_id)))).scalar_one_or_none()
    if not dst:
        dst = M.Inventory(variant_id=payload.variant_id, store_id=payload.to_store_id, quantity=0)
        db.add(dst); await db.flush()
    src.quantity -= payload.quantity
    dst.quantity += payload.quantity
    ref = f"TRF-{uuid.uuid4().hex[:6].upper()}"
    db.add(M.StockMovement(variant_id=payload.variant_id, store_id=payload.from_store_id,
                            type="transfer_out", quantity=payload.quantity,
                            reason=payload.reason or "Transfer", reference=ref, user_id=user.user_id))
    db.add(M.StockMovement(variant_id=payload.variant_id, store_id=payload.to_store_id,
                            type="transfer_in", quantity=payload.quantity,
                            reason=payload.reason or "Transfer", reference=ref, user_id=user.user_id))
    await db.commit()
    return {"ok": True, "reference": ref, "from_qty": src.quantity, "to_qty": dst.quantity}


@api.get("/admin/stock-movements")
async def list_stock_movements(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin), limit: int = 100):
    rows = (await db.execute(select(M.StockMovement).order_by(desc(M.StockMovement.created_at)).limit(limit))).scalars().all()
    out = []
    for m in rows:
        v = (await db.execute(select(M.Variant).where(M.Variant.id == m.variant_id))).scalar_one_or_none()
        p = (await db.execute(select(M.Product).where(M.Product.id == v.product_id))).scalar_one_or_none() if v else None
        s = (await db.execute(select(M.Store).where(M.Store.id == m.store_id))).scalar_one_or_none()
        out.append({
            "id": m.id, "type": m.type, "quantity": m.quantity, "reason": m.reason,
            "reference": m.reference, "store_name": s.name if s else "",
            "product_name": p.name if p else "",
            "variant_label": f"{v.size or ''} / {v.color or ''}" if v else "",
            "created_at": m.created_at.isoformat(),
        })
    return out


# ========== STORES ==========
class StoreIn(BaseModel):
    name: str
    address: Optional[str] = None
    phone: Optional[str] = None
    is_online: bool = False
    active: bool = True


@api.get("/admin/stores")
async def list_stores(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    rows = (await db.execute(select(M.Store).order_by(desc(M.Store.is_online), M.Store.name))).scalars().all()
    return [{"id": s.id, "name": s.name, "address": s.address, "phone": s.phone, "is_online": s.is_online, "active": s.active} for s in rows]


@api.post("/admin/stores")
async def create_store(payload: StoreIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    s = M.Store(**payload.model_dump())
    db.add(s); await db.commit(); await db.refresh(s)
    return {"id": s.id}


@api.put("/admin/stores/{sid}")
async def update_store(sid: str, payload: StoreIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    s = (await db.execute(select(M.Store).where(M.Store.id == sid))).scalar_one_or_none()
    if not s:
        raise HTTPException(404, "Not found")
    for k, v in payload.model_dump().items():
        setattr(s, k, v)
    await db.commit()
    return {"ok": True}


@api.delete("/admin/stores/{sid}")
async def delete_store(sid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    s = (await db.execute(select(M.Store).where(M.Store.id == sid))).scalar_one_or_none()
    if not s:
        raise HTTPException(404, "Not found")
    await db.delete(s); await db.commit()
    return {"ok": True}


# ========== SHIPPING & PAYMENTS ==========
@api.get("/locations")
async def get_locations():
    return {"districts": all_districts(), "by_district": SL_DISTRICT_CITIES}


@api.get("/shipping/quote")
async def shipping_quote(district: Optional[str] = None, city: Optional[str] = None, subtotal: float = 0.0, db: AsyncSession = Depends(get_db)):
    rules = (await db.execute(select(M.ShippingRule).where(M.ShippingRule.active == True).order_by(M.ShippingRule.sort_order))).scalars().all()
    # Most specific (district+city) > district-only > default fallback (district null)
    matched = None
    for r in rules:
        if r.district and r.city and r.district == district and r.city == city:
            matched = r; break
    if not matched:
        for r in rules:
            if r.district and not r.city and r.district == district:
                matched = r; break
    if not matched:
        for r in rules:
            if not r.district and not r.city:
                matched = r; break
    fee = matched.fee if matched else 0.0
    if matched and matched.free_above is not None and subtotal >= matched.free_above:
        fee = 0.0
    return {"fee": fee, "matched_rule_id": matched.id if matched else None,
            "label": matched.label if matched else None}


def _shipping_rule_dict(r):
    return {"id": r.id, "district": r.district, "city": r.city, "fee": r.fee,
            "free_above": r.free_above, "label": r.label, "active": r.active, "sort_order": r.sort_order}


@api.get("/admin/shipping/rules")
async def list_shipping_rules(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    rows = (await db.execute(select(M.ShippingRule).order_by(M.ShippingRule.sort_order, M.ShippingRule.district, M.ShippingRule.city))).scalars().all()
    return [_shipping_rule_dict(r) for r in rows]


class ShippingRuleIn(BaseModel):
    district: Optional[str] = None
    city: Optional[str] = None
    fee: float = 0.0
    free_above: Optional[float] = None
    label: Optional[str] = None
    active: bool = True
    sort_order: int = 0


@api.post("/admin/shipping/rules")
async def create_shipping_rule(payload: ShippingRuleIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    r = M.ShippingRule(**payload.model_dump())
    db.add(r); await db.commit(); await db.refresh(r)
    return _shipping_rule_dict(r)


@api.put("/admin/shipping/rules/{rid}")
async def update_shipping_rule(rid: str, payload: ShippingRuleIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    r = (await db.execute(select(M.ShippingRule).where(M.ShippingRule.id == rid))).scalar_one_or_none()
    if not r:
        raise HTTPException(404, "Not found")
    for k, v in payload.model_dump().items():
        setattr(r, k, v)
    await db.commit()
    return _shipping_rule_dict(r)


@api.delete("/admin/shipping/rules/{rid}")
async def delete_shipping_rule(rid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    r = (await db.execute(select(M.ShippingRule).where(M.ShippingRule.id == rid))).scalar_one_or_none()
    if not r:
        raise HTTPException(404, "Not found")
    await db.delete(r); await db.commit()
    return {"ok": True}


def _payment_dict(p):
    return {"id": p.id, "code": p.code, "label": p.label, "description": p.description,
            "scope": p.scope, "active": p.active, "sort_order": p.sort_order, "config": p.config or {}}


@api.get("/payment-methods")
async def public_payment_methods(scope: str = "online", db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(select(M.PaymentMethod).where(and_(M.PaymentMethod.scope == scope, M.PaymentMethod.active == True)).order_by(M.PaymentMethod.sort_order))).scalars().all()
    # Don't leak provider config publicly
    return [{"id": p.id, "code": p.code, "label": p.label, "description": p.description, "scope": p.scope} for p in rows]


@api.get("/admin/payment-methods")
async def list_payment_methods(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    rows = (await db.execute(select(M.PaymentMethod).order_by(M.PaymentMethod.scope, M.PaymentMethod.sort_order))).scalars().all()
    return [_payment_dict(p) for p in rows]


class PaymentMethodIn(BaseModel):
    code: str
    label: str
    description: Optional[str] = None
    scope: str = "online"
    active: bool = True
    sort_order: int = 0
    config: dict = {}


@api.post("/admin/payment-methods")
async def create_payment_method(payload: PaymentMethodIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    p = M.PaymentMethod(**payload.model_dump())
    db.add(p); await db.commit(); await db.refresh(p)
    return _payment_dict(p)


@api.put("/admin/payment-methods/{pid}")
async def update_payment_method(pid: str, payload: PaymentMethodIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    p = (await db.execute(select(M.PaymentMethod).where(M.PaymentMethod.id == pid))).scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Not found")
    for k, v in payload.model_dump().items():
        setattr(p, k, v)
    await db.commit()
    return _payment_dict(p)


@api.delete("/admin/payment-methods/{pid}")
async def delete_payment_method(pid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    p = (await db.execute(select(M.PaymentMethod).where(M.PaymentMethod.id == pid))).scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Not found")
    await db.delete(p); await db.commit()
    return {"ok": True}


# ========== ORDERS / CHECKOUT ==========
class OrderItemIn(BaseModel):
    variant_id: str
    quantity: int


class CheckoutIn(BaseModel):
    customer_name: Optional[str] = None
    customer_email: Optional[str] = None
    customer_phone: Optional[str] = None
    shipping_address: Optional[str] = None
    shipping_district: Optional[str] = None
    shipping_city: Optional[str] = None
    items: List[OrderItemIn]
    coupon_code: Optional[str] = None
    payment_method: str = "cod"
    notes: Optional[str] = None
    source: str = "online"
    store_id: Optional[str] = None
    cash_tendered: Optional[float] = None  # POS: amount paid in cash
    card_last4: Optional[str] = None        # POS: card terminal last 4
    cash_account_id: Optional[str] = None   # which drawer/account received money
    manual_discount_percent: Optional[float] = None  # POS-only: cashier-entered discount %
    manual_discount_amount: Optional[float] = None   # POS-only: cashier-entered fixed discount


async def _log_notification(db, channel, to, subject, body, order_id=None, status="mocked", provider=None):
    db.add(M.NotificationLog(channel=channel, to_address=to, subject=subject, body=body,
                              related_order=order_id, status=status, provider=provider))


async def _build_order_response(o: M.Order, db: AsyncSession):
    items = (await db.execute(select(M.OrderItem).where(M.OrderItem.order_id == o.id))).scalars().all()
    return {
        "id": o.id, "order_number": o.order_number, "status": o.status,
        "payment_method": o.payment_method, "payment_status": o.payment_status,
        "subtotal": o.subtotal, "discount": o.discount, "coupon_code": o.coupon_code,
        "shipping": o.shipping, "tax": o.tax, "total": o.total,
        "cash_tendered": o.cash_tendered, "cash_change": o.cash_change, "card_last4": o.card_last4,
        "customer_name": o.customer_name, "customer_email": o.customer_email,
        "customer_phone": o.customer_phone, "shipping_address": o.shipping_address,
        "shipping_district": o.shipping_district, "shipping_city": o.shipping_city,
        "store_id": o.store_id,
        "source": o.source, "notes": o.notes, "created_at": o.created_at.isoformat(),
        "items": [{"id": i.id, "product_name": i.product_name, "variant_label": i.variant_label,
                   "unit_price": i.unit_price, "quantity": i.quantity, "subtotal": i.subtotal} for i in items],
    }


async def _resolve_shipping_fee(db, district, city, subtotal):
    rules = (await db.execute(select(M.ShippingRule).where(M.ShippingRule.active == True).order_by(M.ShippingRule.sort_order))).scalars().all()
    matched = None
    for r in rules:
        if r.district and r.city and r.district == district and r.city == city:
            matched = r; break
    if not matched:
        for r in rules:
            if r.district and not r.city and r.district == district:
                matched = r; break
    if not matched:
        for r in rules:
            if not r.district and not r.city:
                matched = r; break
    if not matched:
        return 0.0
    fee = matched.fee
    if matched.free_above is not None and subtotal >= matched.free_above:
        fee = 0.0
    return fee


@api.post("/checkout")
async def checkout(payload: CheckoutIn, request: Request, db: AsyncSession = Depends(get_db)):
    if not payload.items:
        raise HTTPException(400, "Empty cart")
    store = await _ensure_default_store(db)
    store_id = payload.store_id or store.id
    user = await get_current_user_optional(request, db)
    # Normalise phone for SL → +94 E.164 so SMS delivery works downstream.
    if payload.customer_phone:
        payload.customer_phone = normalize_phone_lk(payload.customer_phone)
    customer = None
    if user:
        customer = (await db.execute(select(M.Customer).where(M.Customer.user_id == user.user_id))).scalar_one_or_none()
    if not customer and payload.customer_phone:
        customer = (await db.execute(select(M.Customer).where(M.Customer.phone == payload.customer_phone))).scalar_one_or_none()
    if not customer and payload.customer_email:
        customer = (await db.execute(select(M.Customer).where(M.Customer.email == payload.customer_email))).scalar_one_or_none()
    if not customer:
        customer = M.Customer(user_id=user.user_id if user else None,
                              name=payload.customer_name or "Walk-in",
                              email=payload.customer_email,
                              phone=payload.customer_phone, address=payload.shipping_address,
                              district=payload.shipping_district, city=payload.shipping_city)
        db.add(customer); await db.flush()
    else:
        # Update customer profile with latest info
        if payload.customer_name and not customer.name:
            customer.name = payload.customer_name
        if payload.customer_email and not customer.email:
            customer.email = payload.customer_email
        if payload.customer_phone and not customer.phone:
            customer.phone = payload.customer_phone
        if payload.shipping_district:
            customer.district = payload.shipping_district
        if payload.shipping_city:
            customer.city = payload.shipping_city

    # Active sitewide / category / product discount campaigns auto-apply per item.
    active_discounts = await _select_active_discounts(db)

    subtotal = 0.0
    auto_discount_total = 0.0
    order_items_data = []
    for it in payload.items:
        v = (await db.execute(select(M.Variant).where(M.Variant.id == it.variant_id))).scalar_one_or_none()
        if not v:
            raise HTTPException(400, f"Invalid variant: {it.variant_id}")
        p = (await db.execute(select(M.Product).where(M.Product.id == v.product_id))).scalar_one()
        inv = (await db.execute(select(M.Inventory).where(and_(M.Inventory.variant_id == v.id, M.Inventory.store_id == store_id)))).scalar_one_or_none()
        if not inv or inv.quantity < it.quantity:
            raise HTTPException(400, f"Insufficient stock for {p.name}")
        price = v.price_override if v.price_override is not None else p.base_price
        save_per_unit, applied_d = _best_discount_for(p, price, active_discounts)
        line_save = round(save_per_unit * it.quantity, 2)
        line_subtotal = round(price * it.quantity, 2)
        subtotal += line_subtotal
        auto_discount_total += line_save
        variant_label = f"{v.size or ''} / {v.color or ''}".strip(" /")
        order_items_data.append((v, p, inv, price, it.quantity, line_subtotal, variant_label))

    discount = 0.0
    coupon = None
    if payload.coupon_code:
        coupon = (await db.execute(select(M.Coupon).where(M.Coupon.code == payload.coupon_code.upper()))).scalar_one_or_none()
        if not coupon or not coupon.active:
            raise HTTPException(400, "Invalid coupon")
        if coupon.expires_at:
            exp = coupon.expires_at
            if exp.tzinfo is None:
                exp = exp.replace(tzinfo=timezone.utc)
            if exp < datetime.now(timezone.utc):
                raise HTTPException(400, "Coupon expired")
        if coupon.usage_limit and coupon.used_count >= coupon.usage_limit:
            raise HTTPException(400, "Coupon usage limit reached")
        if subtotal < coupon.min_order:
            raise HTTPException(400, f"Minimum order {coupon.min_order} required")
        # Coupon scope: only discount the eligible portion (products or categories)
        eligible_subtotal = subtotal
        scope = (coupon.scope or "all")
        if scope == "products":
            allowed = set(coupon.scope_product_ids or [])
            eligible_subtotal = sum(line_sub for v, p, inv, price, qty, line_sub, vlabel in order_items_data if p.id in allowed)
        elif scope == "categories":
            allowed = set(coupon.scope_category_ids or [])
            eligible_subtotal = sum(line_sub for v, p, inv, price, qty, line_sub, vlabel in order_items_data if p.category_id in allowed)
        if eligible_subtotal <= 0:
            raise HTTPException(400, "Coupon does not apply to any item in cart")
        discount = round(eligible_subtotal * (coupon.value / 100.0), 2) if coupon.type == "percent" else min(eligible_subtotal, coupon.value)

    if payload.source == "pos":
        shipping_fee = 0.0
    else:
        shipping_fee = await _resolve_shipping_fee(db, payload.shipping_district, payload.shipping_city, subtotal)
    # Auto-applied discount campaigns stack with manual coupon code, but only if both apply.
    discount = round(discount + auto_discount_total, 2)
    # POS may pass a manual discount (percent or fixed) that the cashier punched in.
    pos_extra_discount = 0.0
    if payload.source == "pos":
        if getattr(payload, "manual_discount_percent", None):
            pos_extra_discount = round((subtotal - discount) * (float(payload.manual_discount_percent) / 100.0), 2)
        elif getattr(payload, "manual_discount_amount", None):
            pos_extra_discount = round(min(max(0.0, subtotal - discount), float(payload.manual_discount_amount)), 2)
        discount = round(discount + pos_extra_discount, 2)
    total = max(0.0, subtotal - discount + shipping_fee)

    # Resolve payment status by method.
    # Card / instant methods are auto-paid AND auto-completed (final state).
    # COD / pending gateways stay in pending until manually marked received.
    instant_paid = {"cash", "card_pos", "card", "stripe", "koko", "mintpay", "koko_pos", "mintpay_pos"}
    if payload.payment_method in instant_paid:
        payment_status = "paid"; order_status = "completed"
    else:
        payment_status = "pending"; order_status = "pending"

    # Auto-pick a cash/bank account if the caller didn't supply one (POS source) or for COD
    # (so admins don't need to remember when accounting is enforced).
    target_account_id = payload.cash_account_id
    if not target_account_id:
        wanted_kind = "cash" if payload.payment_method == "cash" else "bank"
        # 1) account bound to the chosen store
        if store_id:
            cand = (await db.execute(select(M.CashAccount).where(and_(
                M.CashAccount.store_id == store_id, M.CashAccount.kind == wanted_kind,
                M.CashAccount.active == True,
            )))).scalars().first()
            if cand:
                target_account_id = cand.id
        # 2) for online orders, account bound to the online store
        if not target_account_id and payload.source == "online":
            online = (await db.execute(select(M.Store).where(M.Store.is_online == True))).scalars().first()
            if online:
                cand = (await db.execute(select(M.CashAccount).where(and_(
                    M.CashAccount.store_id == online.id, M.CashAccount.kind == wanted_kind,
                    M.CashAccount.active == True,
                )))).scalars().first()
                if cand:
                    target_account_id = cand.id
    # Pre-flight: instant-paid orders REQUIRE a destination account so we can credit it.
    # Pending COD orders may proceed without one (admin completes via Cash Received later).
    if payment_status == "paid" and not target_account_id:
        store_label = "the selected store"
        if payload.source == "online":
            store_label = "the online store"
        elif store_id:
            st = (await db.execute(select(M.Store).where(M.Store.id == store_id))).scalar_one_or_none()
            if st: store_label = st.name
        raise HTTPException(400, f"No active cash/bank account configured for {store_label}. Add one in Cash & Bank first.")

    cash_tendered = payload.cash_tendered
    cash_change = None
    if payload.payment_method == "cash" and cash_tendered is not None:
        cash_change = max(0.0, cash_tendered - max(0.0, subtotal - discount + (0.0 if payload.source == "pos" else 0.0)))

    order = M.Order(
        order_number=new_order_number(), customer_id=customer.id,
        customer_name=payload.customer_name, customer_email=payload.customer_email,
        customer_phone=payload.customer_phone, shipping_address=payload.shipping_address,
        shipping_district=payload.shipping_district, shipping_city=payload.shipping_city,
        status=order_status, payment_method=payload.payment_method, payment_status=payment_status,
        subtotal=subtotal, discount=discount, coupon_code=coupon.code if coupon else None,
        shipping=shipping_fee, total=total, store_id=store_id,
        cash_tendered=cash_tendered, cash_change=cash_change,
        card_last4=payload.card_last4, cash_account_id=target_account_id,
        created_by=user.user_id if user else None, source=payload.source, notes=payload.notes,
    )
    db.add(order); await db.flush()
    for v, p, inv, price, qty, line_sub, vlabel in order_items_data:
        db.add(M.OrderItem(order_id=order.id, variant_id=v.id, product_id=p.id,
                            product_name=p.name, variant_label=vlabel, unit_price=price,
                            quantity=qty, subtotal=line_sub))
        inv.quantity = max(0, inv.quantity - qty)
        db.add(M.StockMovement(variant_id=v.id, store_id=store_id, type="sale", quantity=qty,
                                reason=f"Order {order.order_number}", reference=order.order_number,
                                user_id=user.user_id if user else None))
    if coupon:
        coupon.used_count += 1
    customer.total_orders += 1
    customer.total_spent += total
    # Cash ledger: log instant-paid orders to cash account if specified
    if payment_status == "paid" and target_account_id:
        ca = (await db.execute(select(M.CashAccount).where(M.CashAccount.id == target_account_id))).scalar_one_or_none()
        if ca:
            ca.balance += total
            db.add(M.CashLedger(cash_account_id=ca.id, direction="in", amount=total,
                                 source_kind="order", source_id=order.id,
                                 notes=f"Order {order.order_number}",
                                 created_by=user.user_id if user else None))
    # Build receipt URL for SMS
    receipt_path = f"/receipt/{order.order_number}"
    if payload.customer_email:
        await _log_notification(db, "email", payload.customer_email,
                                 f"Order {order.order_number} confirmed",
                                 f"Thank you {payload.customer_name}! Order {order.order_number} total {total:.2f}. Receipt: {receipt_path}",
                                 order.id)
    if payload.customer_phone:
        await _log_notification(db, "sms", payload.customer_phone, "Order Confirmed",
                                 f"Order {order.order_number} confirmed. Total {total:.2f}. Receipt: {receipt_path}",
                                 order.id)
    await db.commit()
    await db.refresh(order)
    return await _build_order_response(order, db)


@api.get("/orders/{order_number}")
async def get_order(order_number: str, db: AsyncSession = Depends(get_db)):
    o = (await db.execute(select(M.Order).where(M.Order.order_number == order_number))).scalar_one_or_none()
    if not o:
        raise HTTPException(404, "Not found")
    return await _build_order_response(o, db)


@api.get("/my/orders")
async def my_orders(user: M.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    customer = (await db.execute(select(M.Customer).where(M.Customer.user_id == user.user_id))).scalar_one_or_none()
    if not customer:
        return []
    rows = (await db.execute(select(M.Order).where(M.Order.customer_id == customer.id).order_by(desc(M.Order.created_at)))).scalars().all()
    return [await _build_order_response(o, db) for o in rows]


@api.get("/admin/orders")
async def admin_orders(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin),
                        status: Optional[str] = None, q: Optional[str] = None,
                        page: int = 1, page_size: int = 50):
    base = select(M.Order)
    if status:
        base = base.where(M.Order.status == status)
    if q:
        base = base.where(or_(
            M.Order.order_number.ilike(f"%{q}%"),
            M.Order.customer_name.ilike(f"%{q}%"),
            M.Order.customer_phone.ilike(f"%{q}%"),
            M.Order.customer_email.ilike(f"%{q}%"),
        ))
    total = (await db.execute(select(func.count()).select_from(base.subquery()))).scalar_one()
    page_size = min(max(1, page_size), 100); page = max(1, page)
    rows = (await db.execute(base.order_by(desc(M.Order.created_at)).offset((page-1)*page_size).limit(page_size))).scalars().all()
    items = [await _build_order_response(o, db) for o in rows]
    return {"items": items, "total": total, "page": page, "page_size": page_size}


class OrderStatusIn(BaseModel):
    status: str


@api.put("/admin/orders/{oid}/status")
async def update_order_status(oid: str, payload: OrderStatusIn, db: AsyncSession = Depends(get_db), user: M.User = Depends(require_admin)):
    o = (await db.execute(select(M.Order).where(M.Order.id == oid))).scalar_one_or_none()
    if not o:
        raise HTTPException(404, "Not found")
    if o.status == "completed":
        raise HTTPException(400, "Completed orders are locked and cannot be changed.")
    new_status = payload.status
    # Pre-paid orders (card / KOKO / Mintpay / stripe / etc.) auto-complete on Delivered
    # and credit the destination BANK account if not already credited.
    prepaid_methods = {"card", "card_pos", "stripe", "koko", "mintpay", "paid"}
    if new_status == "delivered" and o.payment_status == "paid" and o.payment_method in prepaid_methods:
        # If we already wrote to a cash ledger (e.g. POS card), don't double-credit. We mark only when no entry exists.
        existing_ledger = (await db.execute(select(M.CashLedger).where(and_(
            M.CashLedger.source_kind == "order", M.CashLedger.source_id == o.id
        )))).scalars().first()
        if not existing_ledger:
            target = None
            if o.cash_account_id:
                target = (await db.execute(select(M.CashAccount).where(M.CashAccount.id == o.cash_account_id))).scalar_one_or_none()
            # Prefer BANK account on the order's store
            if not target and o.store_id:
                target = (await db.execute(select(M.CashAccount).where(and_(
                    M.CashAccount.store_id == o.store_id, M.CashAccount.kind == "bank", M.CashAccount.active == True,
                )))).scalars().first()
            # Fallback to BANK on online store
            if not target:
                online = (await db.execute(select(M.Store).where(M.Store.is_online == True))).scalars().first()
                if online:
                    target = (await db.execute(select(M.CashAccount).where(and_(
                        M.CashAccount.store_id == online.id, M.CashAccount.kind == "bank", M.CashAccount.active == True,
                    )))).scalars().first()
            if target:
                target.balance += o.total
                o.cash_account_id = target.id
                db.add(M.CashLedger(cash_account_id=target.id, direction="in", amount=o.total,
                                     source_kind="order", source_id=o.id,
                                     notes=f"{o.payment_method.upper()} payout · {o.order_number}",
                                     created_by=user.user_id))
        new_status = "completed"
    o.status = new_status
    if o.customer_email:
        await _log_notification(db, "email", o.customer_email, f"Order {o.order_number} {new_status}",
                                 f"Your order {o.order_number} status: {new_status}.", o.id)
    await db.commit()
    return {"ok": True, "status": o.status}


@api.post("/admin/orders/{oid}/cash-received")
async def mark_cash_received(oid: str, db: AsyncSession = Depends(get_db), user: M.User = Depends(require_admin)):
    o = (await db.execute(select(M.Order).where(M.Order.id == oid))).scalar_one_or_none()
    if not o:
        raise HTTPException(404, "Not found")
    if o.status == "completed":
        raise HTTPException(400, "Order already completed.")
    # COD cash received → goes to BANK account of the store (delivery person banks the cash).
    # Order of preference:
    #   1) The cash_account_id already attached to the order (usually nothing for COD)
    #   2) BANK account bound to the order's store
    #   3) BANK account on the online store
    #   4) CASH account on the online store (last-resort fallback if no bank account exists)
    target = None
    if o.cash_account_id:
        target = (await db.execute(select(M.CashAccount).where(M.CashAccount.id == o.cash_account_id))).scalar_one_or_none()
    if not target and o.store_id:
        target = (await db.execute(select(M.CashAccount).where(and_(
            M.CashAccount.store_id == o.store_id, M.CashAccount.kind == "bank", M.CashAccount.active == True
        )))).scalars().first()
    online = (await db.execute(select(M.Store).where(M.Store.is_online == True))).scalars().first()
    if not target and online:
        target = (await db.execute(select(M.CashAccount).where(and_(
            M.CashAccount.store_id == online.id, M.CashAccount.kind == "bank", M.CashAccount.active == True
        )))).scalars().first()
    if not target and online:
        target = (await db.execute(select(M.CashAccount).where(and_(
            M.CashAccount.store_id == online.id, M.CashAccount.kind == "cash", M.CashAccount.active == True
        )))).scalars().first()
    if not target:
        raise HTTPException(400, "No active bank or cash account configured for the online store. Add one in Cash & Bank first.")
    o.payment_status = "paid"
    o.status = "completed"
    o.cash_account_id = target.id
    target.balance += o.total
    db.add(M.CashLedger(cash_account_id=target.id, direction="in", amount=o.total,
                         source_kind="order", source_id=o.id,
                         notes=f"COD banked · {o.order_number}", created_by=user.user_id))
    if o.customer_email:
        await _log_notification(db, "email", o.customer_email, f"Order {o.order_number} completed",
                                 f"Cash received. Order {o.order_number} is now complete. Thank you.", o.id)
    await db.commit()
    return {"ok": True, "status": o.status, "payment_status": o.payment_status,
            "credited_account_id": target.id, "credited_account_name": target.name}


# ========== ORDER STATS (sidebar badge) ==========
@api.get("/admin/orders/stats")
async def order_stats(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    pending = (await db.execute(select(func.count()).select_from(
        select(M.Order).where(M.Order.status == "pending").subquery()))).scalar_one()
    processing = (await db.execute(select(func.count()).select_from(
        select(M.Order).where(M.Order.status == "processing").subquery()))).scalar_one()
    return {"pending": pending or 0, "processing": processing or 0}


# ========== CUSTOMERS ==========
@api.get("/admin/customers")
async def list_customers(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin), q: Optional[str] = None, limit: int = 200):
    query = select(M.Customer).order_by(desc(M.Customer.created_at))
    if q:
        # Search by name, phone, email, or order_number
        # First check if a matching order_number exists
        order_match = (await db.execute(select(M.Order).where(M.Order.order_number.ilike(f"%{q}%")))).scalars().first()
        if order_match and order_match.customer_id:
            query = query.where(M.Customer.id == order_match.customer_id)
        else:
            query = query.where(or_(
                M.Customer.name.ilike(f"%{q}%"),
                M.Customer.phone.ilike(f"%{q}%"),
                M.Customer.email.ilike(f"%{q}%"),
            ))
    rows = (await db.execute(query.limit(limit))).scalars().all()
    return [{"id": c.id, "name": c.name, "email": c.email, "phone": c.phone, "address": c.address,
             "district": c.district, "city": c.city,
             "notes": c.notes, "total_orders": c.total_orders, "total_spent": c.total_spent,
             "created_at": c.created_at.isoformat()} for c in rows]


class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    district: Optional[str] = None
    city: Optional[str] = None
    notes: Optional[str] = None


class CustomerCreate(BaseModel):
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    district: Optional[str] = None
    city: Optional[str] = None


@api.post("/admin/customers")
async def create_customer(payload: CustomerCreate, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    # Auto-deduplicate by phone or email
    existing = None
    if payload.phone:
        existing = (await db.execute(select(M.Customer).where(M.Customer.phone == payload.phone))).scalar_one_or_none()
    if not existing and payload.email:
        existing = (await db.execute(select(M.Customer).where(M.Customer.email == payload.email))).scalar_one_or_none()
    if existing:
        return {"id": existing.id, "name": existing.name, "phone": existing.phone, "email": existing.email,
                "address": existing.address, "district": existing.district, "city": existing.city, "deduplicated": True}
    c = M.Customer(**payload.model_dump())
    db.add(c); await db.commit(); await db.refresh(c)
    return {"id": c.id, "name": c.name, "phone": c.phone, "email": c.email,
            "address": c.address, "district": c.district, "city": c.city}


@api.put("/admin/customers/{cid}")
async def update_customer(cid: str, payload: CustomerUpdate, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    c = (await db.execute(select(M.Customer).where(M.Customer.id == cid))).scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Not found")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(c, k, v)
    await db.commit()
    return {"ok": True}


# Logged-in customer fetch own profile (for autofill)
@api.get("/my/profile")
async def my_profile(user: M.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    customer = (await db.execute(select(M.Customer).where(M.Customer.user_id == user.user_id))).scalar_one_or_none()
    if not customer:
        return {"name": user.name, "email": user.email, "phone": user.phone}
    return {"name": customer.name, "email": customer.email, "phone": customer.phone,
            "address": customer.address, "district": customer.district, "city": customer.city}


# ========== COUPONS ==========
class CouponIn(BaseModel):
    code: str
    type: str = "percent"
    value: float
    min_order: float = 0.0
    usage_limit: int = 0
    active: bool = True
    expires_at: Optional[datetime] = None
    scope: str = "all"  # all, products, categories
    scope_product_ids: Optional[List[str]] = None
    scope_category_ids: Optional[List[str]] = None


def _coupon_dict(c):
    return {"id": c.id, "code": c.code, "type": c.type, "value": c.value, "min_order": c.min_order,
            "usage_limit": c.usage_limit, "used_count": c.used_count, "active": c.active,
            "expires_at": c.expires_at.isoformat() if c.expires_at else None,
            "scope": c.scope or "all",
            "scope_product_ids": c.scope_product_ids or [],
            "scope_category_ids": c.scope_category_ids or []}


@api.get("/admin/coupons")
async def list_coupons(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin),
                       q: Optional[str] = None, page: int = 1, page_size: int = 50):
    base = select(M.Coupon)
    if q:
        base = base.where(M.Coupon.code.ilike(f"%{q}%"))
    total = (await db.execute(select(func.count()).select_from(base.subquery()))).scalar_one()
    page_size = min(max(1, page_size), 100); page = max(1, page)
    rows = (await db.execute(base.order_by(desc(M.Coupon.created_at)).offset((page-1)*page_size).limit(page_size))).scalars().all()
    items = [_coupon_dict(c) for c in rows]
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@api.post("/admin/coupons")
async def create_coupon(payload: CouponIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    c = M.Coupon(**payload.model_dump()); c.code = c.code.upper()
    db.add(c); await db.commit(); await db.refresh(c)
    return {"id": c.id}


@api.put("/admin/coupons/{cid}")
async def update_coupon(cid: str, payload: CouponIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    c = (await db.execute(select(M.Coupon).where(M.Coupon.id == cid))).scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Not found")
    for k, v in payload.model_dump().items():
        setattr(c, k, v)
    c.code = c.code.upper()
    await db.commit()
    return {"ok": True}


@api.delete("/admin/coupons/{cid}")
async def delete_coupon(cid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    c = (await db.execute(select(M.Coupon).where(M.Coupon.id == cid))).scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Not found")
    await db.delete(c); await db.commit()
    return {"ok": True}


@api.get("/coupons/validate/{code}")
async def validate_coupon(code: str, db: AsyncSession = Depends(get_db)):
    c = (await db.execute(select(M.Coupon).where(M.Coupon.code == code.upper()))).scalar_one_or_none()
    if not c or not c.active:
        raise HTTPException(404, "Invalid coupon")
    return {"code": c.code, "type": c.type, "value": c.value, "min_order": c.min_order}


# ========== EXPENSES ==========
class ExpenseIn(BaseModel):
    category: str
    amount: float
    description: Optional[str] = None
    expense_date: Optional[datetime] = None
    store_id: Optional[str] = None
    method: str = "cash"
    cash_account_id: Optional[str] = None


@api.get("/admin/expenses")
async def list_expenses(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin),
                         q: Optional[str] = None, page: int = 1, page_size: int = 50,
                         store_id: Optional[str] = None):
    base = select(M.Expense)
    if q:
        base = base.where(or_(M.Expense.category.ilike(f"%{q}%"), M.Expense.description.ilike(f"%{q}%")))
    if store_id:
        base = base.where(M.Expense.store_id == store_id)
    total = (await db.execute(select(func.count()).select_from(base.subquery()))).scalar_one()
    page_size = min(max(1, page_size), 100); page = max(1, page)
    rows = (await db.execute(base.order_by(desc(M.Expense.expense_date)).offset((page-1)*page_size).limit(page_size))).scalars().all()
    items = [{"id": e.id, "category": e.category, "amount": e.amount, "description": e.description,
              "store_id": e.store_id, "method": e.method, "cash_account_id": e.cash_account_id,
              "expense_date": e.expense_date.isoformat()} for e in rows]
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@api.post("/admin/expenses")
async def create_expense(payload: ExpenseIn, db: AsyncSession = Depends(get_db), user: M.User = Depends(require_perm("manual_inc_exp"))):
    e = M.Expense(**payload.model_dump(exclude_unset=True))
    e.created_by = user.user_id
    if not e.expense_date:
        e.expense_date = datetime.now(timezone.utc)
    db.add(e); await db.flush()
    # Cash ledger if assigned to an account
    if payload.cash_account_id:
        ca = (await db.execute(select(M.CashAccount).where(M.CashAccount.id == payload.cash_account_id))).scalar_one_or_none()
        if ca:
            ca.balance -= payload.amount
            db.add(M.CashLedger(cash_account_id=ca.id, direction="out", amount=payload.amount,
                                source_kind="expense", source_id=e.id,
                                notes=f"{e.category}: {e.description or ''}",
                                created_by=user.user_id))
    await db.commit(); await db.refresh(e)
    return {"id": e.id}


@api.delete("/admin/expenses/{eid}")
async def delete_expense(eid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    e = (await db.execute(select(M.Expense).where(M.Expense.id == eid))).scalar_one_or_none()
    if not e:
        raise HTTPException(404, "Not found")
    await db.delete(e); await db.commit()
    return {"ok": True}


# ========== PAYROLL ==========
class PayrollIn(BaseModel):
    staff_user_id: str
    month: int
    year: int
    base_salary: float
    bonus: float = 0.0
    deduction: float = 0.0
    status: str = "pending"


@api.get("/admin/payroll")
async def list_payroll(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    rows = (await db.execute(select(M.Payroll).order_by(desc(M.Payroll.created_at)))).scalars().all()
    out = []
    for p in rows:
        u = (await db.execute(select(M.User).where(M.User.user_id == p.staff_user_id))).scalar_one_or_none()
        out.append({"id": p.id, "staff_user_id": p.staff_user_id, "staff_name": u.name if u else "",
                    "month": p.month, "year": p.year, "base_salary": p.base_salary, "bonus": p.bonus,
                    "deduction": p.deduction, "net": p.net, "status": p.status,
                    "paid_date": p.paid_date.isoformat() if p.paid_date else None})
    return out


@api.post("/admin/payroll")
async def create_payroll(payload: PayrollIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    net = payload.base_salary + payload.bonus - payload.deduction
    p = M.Payroll(**payload.model_dump(), net=net)
    if payload.status == "paid":
        p.paid_date = datetime.now(timezone.utc)
    db.add(p); await db.commit(); await db.refresh(p)
    return {"id": p.id}


@api.put("/admin/payroll/{pid}/pay")
async def mark_payroll_paid(pid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    p = (await db.execute(select(M.Payroll).where(M.Payroll.id == pid))).scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Not found")
    p.status = "paid"; p.paid_date = datetime.now(timezone.utc)
    await db.commit()
    return {"ok": True}


# ========== STAFF ==========
class StaffIn(BaseModel):
    email: EmailStr
    name: str
    phone: Optional[str] = None
    role: str
    base_salary: Optional[float] = None
    active: bool = True
    password: Optional[str] = None  # Optional initial password
    permissions: Optional[dict] = None


@api.get("/admin/staff")
async def list_staff(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin), q: Optional[str] = None):
    query = select(M.User).where(M.User.role != "customer").order_by(desc(M.User.created_at))
    if q:
        query = query.where(or_(M.User.name.ilike(f"%{q}%"), M.User.email.ilike(f"%{q}%")))
    rows = (await db.execute(query)).scalars().all()
    return [{"user_id": u.user_id, "email": u.email, "name": u.name, "phone": u.phone,
             "role": u.role, "base_salary": u.base_salary, "active": u.active,
             "auth_provider": u.auth_provider, "permissions": u.permissions or {},
             "created_at": u.created_at.isoformat()} for u in rows]


@api.post("/admin/staff")
async def create_staff(payload: StaffIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_roles("super_admin"))):
    # Default to "all permissions OFF" — admin must explicitly grant.
    perms = payload.permissions or {p: False for p in ALL_PERMISSIONS}
    existing = (await db.execute(select(M.User).where(M.User.email == payload.email.lower()))).scalar_one_or_none()
    if existing:
        existing.role = payload.role; existing.name = payload.name
        existing.phone = payload.phone; existing.base_salary = payload.base_salary
        existing.active = payload.active; existing.permissions = perms
        if payload.password:
            existing.password_hash = hash_password(payload.password)
            existing.auth_provider = "password"
        await db.commit()
        return {"user_id": existing.user_id}
    u = M.User(user_id=f"user_{uuid.uuid4().hex[:12]}", email=payload.email.lower(), name=payload.name,
               phone=payload.phone, role=payload.role, base_salary=payload.base_salary,
               active=payload.active, auth_provider="password", permissions=perms,
               password_hash=hash_password(payload.password) if payload.password else None)
    db.add(u); await db.commit(); await db.refresh(u)
    return {"user_id": u.user_id}


@api.put("/admin/staff/{uid}")
async def update_staff(uid: str, payload: StaffIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_roles("super_admin"))):
    u = (await db.execute(select(M.User).where(M.User.user_id == uid))).scalar_one_or_none()
    if not u:
        raise HTTPException(404, "Not found")
    u.email = payload.email.lower(); u.name = payload.name; u.phone = payload.phone
    u.role = payload.role; u.base_salary = payload.base_salary; u.active = payload.active
    if payload.permissions is not None:
        u.permissions = payload.permissions
    if payload.password:
        u.password_hash = hash_password(payload.password)
        u.auth_provider = "password"
    await db.commit()
    return {"ok": True}


@api.delete("/admin/staff/{uid}")
async def delete_staff(uid: str, db: AsyncSession = Depends(get_db), current: M.User = Depends(require_roles("super_admin"))):
    if uid == current.user_id:
        raise HTTPException(400, "Cannot delete yourself")
    u = (await db.execute(select(M.User).where(M.User.user_id == uid))).scalar_one_or_none()
    if not u:
        raise HTTPException(404, "Not found")
    u.active = False; u.role = "customer"
    await db.commit()
    return {"ok": True}


# ========== REPORTS / DASHBOARD ==========
@api.get("/admin/dashboard")
async def dashboard(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    now = datetime.now(timezone.utc); d30 = now - timedelta(days=30); d14 = now - timedelta(days=14)
    orders_30 = (await db.execute(select(M.Order).where(M.Order.created_at >= d30))).scalars().all()
    total_revenue = sum(o.total for o in orders_30 if o.payment_status == "paid")
    customer_count = (await db.execute(select(func.count(M.Customer.id)))).scalar_one()
    low_rows = (await db.execute(select(M.Inventory))).scalars().all()
    low_stock = [x for x in low_rows if x.quantity <= x.low_stock_threshold]
    daily_orders = (await db.execute(select(M.Order).where(M.Order.created_at >= d14))).scalars().all()
    daily = {}
    for o in daily_orders:
        if o.payment_status != "paid":
            continue
        d = o.created_at.strftime("%Y-%m-%d")
        daily[d] = daily.get(d, 0) + o.total
    sales_chart = [{"date": k, "revenue": round(v, 2)} for k, v in sorted(daily.items())]
    items = (await db.execute(select(M.OrderItem))).scalars().all()
    top_map = {}
    for i in items:
        top_map[i.product_name] = top_map.get(i.product_name, 0) + i.quantity
    top_products = sorted([{"name": k, "qty": v} for k, v in top_map.items()], key=lambda x: -x["qty"])[:5]
    status_map = {}
    for o in orders_30:
        status_map[o.status] = status_map.get(o.status, 0) + 1
    return {"total_revenue": round(total_revenue, 2), "total_orders": len(orders_30),
            "customer_count": customer_count, "low_stock_count": len(low_stock),
            "sales_chart": sales_chart, "top_products": top_products,
            "status_breakdown": [{"status": k, "count": v} for k, v in status_map.items()]}


@api.get("/admin/reports/sales")
async def sales_report(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin), days: int = 30):
    since = datetime.now(timezone.utc) - timedelta(days=days)
    orders = (await db.execute(select(M.Order).where(M.Order.created_at >= since))).scalars().all()
    by_day = {}; by_channel = {}; total_paid = 0.0
    for o in orders:
        d = o.created_at.strftime("%Y-%m-%d")
        by_day[d] = by_day.get(d, 0.0) + (o.total if o.payment_status == "paid" else 0)
        by_channel[o.source] = by_channel.get(o.source, 0.0) + (o.total if o.payment_status == "paid" else 0)
        if o.payment_status == "paid":
            total_paid += o.total
    exp = (await db.execute(select(M.Expense).where(M.Expense.expense_date >= since))).scalars().all()
    total_expense = sum(e.amount for e in exp)
    return {"total_paid_revenue": round(total_paid, 2), "total_expenses": round(total_expense, 2),
            "profit": round(total_paid - total_expense, 2),
            "by_day": [{"date": k, "revenue": round(v, 2)} for k, v in sorted(by_day.items())],
            "by_channel": [{"channel": k, "revenue": round(v, 2)} for k, v in by_channel.items()]}


# ========== MARKETING ==========
class CampaignIn(BaseModel):
    name: str; channel: str; status: str = "draft"
    spend: float = 0.0; revenue: float = 0.0; reach: int = 0
    clicks: int = 0; conversions: int = 0
    start_date: Optional[datetime] = None; end_date: Optional[datetime] = None


@api.get("/admin/marketing/campaigns")
async def list_campaigns(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    rows = (await db.execute(select(M.MarketingCampaign).order_by(desc(M.MarketingCampaign.created_at)))).scalars().all()
    return [{"id": c.id, "name": c.name, "channel": c.channel, "status": c.status,
             "spend": c.spend, "revenue": c.revenue, "reach": c.reach, "clicks": c.clicks,
             "conversions": c.conversions,
             "roi": round((c.revenue - c.spend) / c.spend * 100, 2) if c.spend else 0} for c in rows]


@api.post("/admin/marketing/campaigns")
async def create_campaign(payload: CampaignIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    c = M.MarketingCampaign(**payload.model_dump())
    db.add(c); await db.commit(); await db.refresh(c)
    return {"id": c.id}


@api.put("/admin/marketing/campaigns/{cid}")
async def update_campaign(cid: str, payload: CampaignIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    c = (await db.execute(select(M.MarketingCampaign).where(M.MarketingCampaign.id == cid))).scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Not found")
    for k, v in payload.model_dump().items():
        setattr(c, k, v)
    await db.commit()
    return {"ok": True}


@api.delete("/admin/marketing/campaigns/{cid}")
async def delete_campaign(cid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    c = (await db.execute(select(M.MarketingCampaign).where(M.MarketingCampaign.id == cid))).scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Not found")
    await db.delete(c); await db.commit()
    return {"ok": True}


@api.get("/admin/notifications")
async def list_notifications(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin), limit: int = 50):
    rows = (await db.execute(select(M.NotificationLog).order_by(desc(M.NotificationLog.created_at)).limit(limit))).scalars().all()
    return [{"id": n.id, "channel": n.channel, "to": n.to_address, "subject": n.subject,
             "body": n.body, "related_order": n.related_order, "status": n.status,
             "provider": n.provider, "created_at": n.created_at.isoformat()} for n in rows]


# ========== PAGES & PAGE BUILDER ==========
class SectionIn(BaseModel):
    section_type: str
    sort_order: int = 0
    visible: bool = True
    config: dict = {}


class SectionUpdate(BaseModel):
    sort_order: Optional[int] = None
    visible: Optional[bool] = None
    config: Optional[dict] = None


class ThemeIn(BaseModel):
    config: dict


def _section_to_dict(s):
    return {"id": s.id, "page": s.page, "section_type": s.section_type,
            "sort_order": s.sort_order, "visible": s.visible, "config": s.config or {}}


def _page_to_dict(p):
    return {"id": p.id, "slug": p.slug, "title": p.title, "is_system": p.is_system,
            "show_in_header_menu": p.show_in_header_menu, "sort_order": p.sort_order,
            "visible": p.visible}


@api.get("/page/{page}")
async def get_page(page: str, db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(
        select(M.PageSection).where(and_(M.PageSection.page == page, M.PageSection.visible == True)).order_by(M.PageSection.sort_order)
    )).scalars().all()
    theme = (await db.execute(select(M.ThemeSetting).where(M.ThemeSetting.id == "default"))).scalar_one_or_none()
    meta = (await db.execute(select(M.CustomPage).where(M.CustomPage.slug == page))).scalar_one_or_none()
    return {"sections": [_section_to_dict(s) for s in rows],
            "theme": (theme.config if theme else DEFAULT_THEME),
            "meta": _page_to_dict(meta) if meta else None}


@api.get("/pages")
async def list_public_pages(db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(select(M.CustomPage).where(and_(
        M.CustomPage.visible == True, M.CustomPage.is_system == False
    )).order_by(M.CustomPage.sort_order, M.CustomPage.title))).scalars().all()
    return [_page_to_dict(p) for p in rows]


@api.get("/admin/pages")
async def admin_list_pages(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    rows = (await db.execute(select(M.CustomPage).order_by(M.CustomPage.is_system.desc(), M.CustomPage.sort_order, M.CustomPage.title))).scalars().all()
    return [_page_to_dict(p) for p in rows]


class PageIn(BaseModel):
    slug: Optional[str] = None
    title: str
    show_in_header_menu: bool = False
    sort_order: int = 0
    visible: bool = True


@api.post("/admin/pages")
async def create_page(payload: PageIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    slug = (payload.slug and slugify(payload.slug)) or slugify(payload.title)
    if (await db.execute(select(M.CustomPage).where(M.CustomPage.slug == slug))).scalar_one_or_none():
        slug = slug + "-" + uuid.uuid4().hex[:4]
    p = M.CustomPage(slug=slug, title=payload.title, show_in_header_menu=payload.show_in_header_menu,
                      sort_order=payload.sort_order, visible=payload.visible, is_system=False)
    db.add(p); await db.commit(); await db.refresh(p)
    # Seed an initial empty heading_text section
    db.add(M.PageSection(page=slug, section_type="custom", sort_order=0, visible=True,
                          config={"block_type": "heading_text", "eyebrow": "", "heading": payload.title,
                                  "text": "Edit this page from Page Builder.", "alignment": "left",
                                  "max_width": "narrow", "padding": "lg"}))
    await db.commit()
    return _page_to_dict(p)


@api.put("/admin/pages/{pid}")
async def update_page(pid: str, payload: PageIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    p = (await db.execute(select(M.CustomPage).where(M.CustomPage.id == pid))).scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Not found")
    p.title = payload.title; p.show_in_header_menu = payload.show_in_header_menu
    p.sort_order = payload.sort_order; p.visible = payload.visible
    await db.commit()
    return _page_to_dict(p)


@api.delete("/admin/pages/{pid}")
async def delete_page(pid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    p = (await db.execute(select(M.CustomPage).where(M.CustomPage.id == pid))).scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Not found")
    if p.is_system:
        raise HTTPException(400, "Cannot delete system page")
    # delete sections too
    for s in (await db.execute(select(M.PageSection).where(M.PageSection.page == p.slug))).scalars().all():
        await db.delete(s)
    await db.delete(p); await db.commit()
    return {"ok": True}


@api.get("/admin/page/{page}")
async def admin_get_page(page: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    rows = (await db.execute(
        select(M.PageSection).where(M.PageSection.page == page).order_by(M.PageSection.sort_order)
    )).scalars().all()
    return {"sections": [_section_to_dict(s) for s in rows]}


@api.post("/admin/page/{page}/sections")
async def add_section(page: str, payload: SectionIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    if payload.sort_order is None or payload.sort_order == 0:
        max_order = (await db.execute(select(func.max(M.PageSection.sort_order)).where(M.PageSection.page == page))).scalar_one() or 0
        payload.sort_order = int(max_order) + 10
    s = M.PageSection(page=page, section_type=payload.section_type, sort_order=payload.sort_order,
                      visible=payload.visible, config=payload.config)
    db.add(s); await db.commit(); await db.refresh(s)
    return _section_to_dict(s)


@api.put("/admin/page/sections/{sid}")
async def update_section(sid: str, payload: SectionUpdate, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    s = (await db.execute(select(M.PageSection).where(M.PageSection.id == sid))).scalar_one_or_none()
    if not s:
        raise HTTPException(404, "Not found")
    if payload.sort_order is not None:
        s.sort_order = payload.sort_order
    if payload.visible is not None:
        s.visible = payload.visible
    if payload.config is not None:
        s.config = payload.config
    s.updated_at = datetime.now(timezone.utc)
    await db.commit()
    return _section_to_dict(s)


@api.delete("/admin/page/sections/{sid}")
async def delete_section(sid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    s = (await db.execute(select(M.PageSection).where(M.PageSection.id == sid))).scalar_one_or_none()
    if not s:
        raise HTTPException(404, "Not found")
    await db.delete(s); await db.commit()
    return {"ok": True}


class ReorderIn(BaseModel):
    ids: list[str]


@api.post("/admin/page/{page}/reorder")
async def reorder_sections(page: str, payload: ReorderIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    for i, sid in enumerate(payload.ids):
        s = (await db.execute(select(M.PageSection).where(and_(M.PageSection.id == sid, M.PageSection.page == page)))).scalar_one_or_none()
        if s:
            s.sort_order = i * 10
    await db.commit()
    return {"ok": True}


@api.get("/theme")
async def get_theme(db: AsyncSession = Depends(get_db)):
    theme = (await db.execute(select(M.ThemeSetting).where(M.ThemeSetting.id == "default"))).scalar_one_or_none()
    return theme.config if theme else DEFAULT_THEME


@api.put("/admin/theme")
async def update_theme(payload: ThemeIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    theme = (await db.execute(select(M.ThemeSetting).where(M.ThemeSetting.id == "default"))).scalar_one_or_none()
    if not theme:
        theme = M.ThemeSetting(id="default", config=payload.config); db.add(theme)
    else:
        theme.config = payload.config; theme.updated_at = datetime.now(timezone.utc)
    await db.commit()
    return theme.config


# ---- Media ----
class MediaUpload(BaseModel):
    data_base64: str
    mime_type: str = "image/png"
    filename: Optional[str] = None


@api.post("/admin/media")
async def upload_media(payload: MediaUpload, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    m = M.Media(data_base64=payload.data_base64, mime_type=payload.mime_type, filename=payload.filename)
    db.add(m); await db.commit(); await db.refresh(m)
    return {"id": m.id, "url": f"/api/media/{m.id}", "mime_type": m.mime_type}


@api.get("/media/{mid}")
async def stream_media(mid: str, db: AsyncSession = Depends(get_db)):
    row = (await db.execute(select(M.Media.data_base64, M.Media.mime_type).where(M.Media.id == mid))).first()
    if not row:
        raise HTTPException(404, "Not found")
    try:
        raw = _b64.b64decode(row.data_base64)
    except Exception:
        raise HTTPException(500, "Corrupt")
    return FastResponse(content=raw, media_type=row.mime_type or "image/png",
                        headers={"Cache-Control": "public, max-age=31536000, immutable"})


@api.delete("/admin/media/{mid}")
async def delete_media(mid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    m = (await db.execute(select(M.Media).where(M.Media.id == mid))).scalar_one_or_none()
    if not m:
        raise HTTPException(404, "Not found")
    await db.delete(m); await db.commit()
    return {"ok": True}


# ========== HEALTH ==========
@api.get("/")
async def root():
    return {"app": "ERP Storefront", "status": "ok"}


# ========== SUPPLIERS ==========
class SupplierIn(BaseModel):
    name: str
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None
    active: bool = True


def _supplier_dict(s):
    return {"id": s.id, "name": s.name, "contact_person": s.contact_person, "phone": s.phone,
            "email": s.email, "address": s.address, "notes": s.notes,
            "balance_owed": s.balance_owed, "total_purchases": s.total_purchases, "total_paid": s.total_paid,
            "active": s.active, "created_at": s.created_at.isoformat()}


@api.get("/admin/suppliers")
async def list_suppliers(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin),
                          q: Optional[str] = None, page: int = 1, page_size: int = 50):
    base = select(M.Supplier)
    if q:
        base = base.where(or_(M.Supplier.name.ilike(f"%{q}%"), M.Supplier.phone.ilike(f"%{q}%"),
                              M.Supplier.email.ilike(f"%{q}%")))
    total = (await db.execute(select(func.count()).select_from(base.subquery()))).scalar_one()
    page_size = min(max(1, page_size), 100); page = max(1, page)
    rows = (await db.execute(base.order_by(desc(M.Supplier.created_at)).offset((page-1)*page_size).limit(page_size))).scalars().all()
    return {"items": [_supplier_dict(s) for s in rows], "total": total, "page": page, "page_size": page_size}


@api.post("/admin/suppliers")
async def create_supplier(payload: SupplierIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    s = M.Supplier(**payload.model_dump())
    db.add(s); await db.commit(); await db.refresh(s)
    return _supplier_dict(s)


@api.put("/admin/suppliers/{sid}")
async def update_supplier(sid: str, payload: SupplierIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    s = (await db.execute(select(M.Supplier).where(M.Supplier.id == sid))).scalar_one_or_none()
    if not s:
        raise HTTPException(404, "Not found")
    for k, v in payload.model_dump().items():
        setattr(s, k, v)
    await db.commit()
    return _supplier_dict(s)


@api.delete("/admin/suppliers/{sid}")
async def delete_supplier(sid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    s = (await db.execute(select(M.Supplier).where(M.Supplier.id == sid))).scalar_one_or_none()
    if not s:
        raise HTTPException(404, "Not found")
    await db.delete(s); await db.commit()
    return {"ok": True}


class SupplierInvoiceIn(BaseModel):
    supplier_id: str
    reference: Optional[str] = None
    amount: float
    notes: Optional[str] = None
    invoice_date: Optional[datetime] = None


@api.get("/admin/suppliers/{sid}/invoices")
async def list_supplier_invoices(sid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    rows = (await db.execute(select(M.SupplierInvoice).where(M.SupplierInvoice.supplier_id == sid)
                              .order_by(desc(M.SupplierInvoice.invoice_date)))).scalars().all()
    return [{"id": r.id, "reference": r.reference, "amount": r.amount, "paid": r.paid, "balance": r.amount - r.paid,
             "notes": r.notes, "invoice_date": r.invoice_date.isoformat()} for r in rows]


@api.post("/admin/supplier-invoices")
async def create_supplier_invoice(payload: SupplierInvoiceIn, db: AsyncSession = Depends(get_db), user: M.User = Depends(require_admin)):
    s = (await db.execute(select(M.Supplier).where(M.Supplier.id == payload.supplier_id))).scalar_one_or_none()
    if not s:
        raise HTTPException(404, "Supplier not found")
    inv = M.SupplierInvoice(supplier_id=payload.supplier_id, reference=payload.reference,
                             amount=payload.amount, notes=payload.notes,
                             invoice_date=payload.invoice_date or datetime.now(timezone.utc),
                             created_by=user.user_id)
    db.add(inv)
    s.total_purchases += payload.amount
    s.balance_owed += payload.amount
    await db.commit(); await db.refresh(inv)
    return {"id": inv.id}


class SupplierPayIn(BaseModel):
    supplier_id: str
    invoice_id: Optional[str] = None
    amount: float
    method: str = "cash"
    cash_account_id: Optional[str] = None
    notes: Optional[str] = None


@api.post("/admin/supplier-payments")
async def pay_supplier(payload: SupplierPayIn, db: AsyncSession = Depends(get_db), user: M.User = Depends(require_admin)):
    s = (await db.execute(select(M.Supplier).where(M.Supplier.id == payload.supplier_id))).scalar_one_or_none()
    if not s:
        raise HTTPException(404, "Supplier not found")
    pay = M.SupplierPayment(supplier_id=payload.supplier_id, invoice_id=payload.invoice_id,
                             amount=payload.amount, method=payload.method,
                             cash_account_id=payload.cash_account_id, notes=payload.notes,
                             created_by=user.user_id)
    db.add(pay)
    s.total_paid += payload.amount
    s.balance_owed = max(0.0, s.balance_owed - payload.amount)
    if payload.invoice_id:
        inv = (await db.execute(select(M.SupplierInvoice).where(M.SupplierInvoice.id == payload.invoice_id))).scalar_one_or_none()
        if inv:
            inv.paid = min(inv.amount, inv.paid + payload.amount)
    if payload.cash_account_id:
        ca = (await db.execute(select(M.CashAccount).where(M.CashAccount.id == payload.cash_account_id))).scalar_one_or_none()
        if ca:
            ca.balance -= payload.amount
            db.add(M.CashLedger(cash_account_id=ca.id, direction="out", amount=payload.amount,
                                source_kind="supplier", source_id=pay.id,
                                notes=f"Paid {s.name}", created_by=user.user_id))
    await db.commit(); await db.refresh(pay)
    return {"id": pay.id, "balance_owed": s.balance_owed}


# ========== INCOME ==========
class IncomeIn(BaseModel):
    category: str
    amount: float
    description: Optional[str] = None
    income_date: Optional[datetime] = None
    store_id: Optional[str] = None
    method: str = "cash"
    cash_account_id: Optional[str] = None


@api.get("/admin/income")
async def list_income(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin),
                       q: Optional[str] = None, page: int = 1, page_size: int = 50,
                       store_id: Optional[str] = None):
    base = select(M.Income)
    if q:
        base = base.where(or_(M.Income.category.ilike(f"%{q}%"), M.Income.description.ilike(f"%{q}%")))
    if store_id:
        base = base.where(M.Income.store_id == store_id)
    total = (await db.execute(select(func.count()).select_from(base.subquery()))).scalar_one()
    page_size = min(max(1, page_size), 100); page = max(1, page)
    rows = (await db.execute(base.order_by(desc(M.Income.income_date)).offset((page-1)*page_size).limit(page_size))).scalars().all()
    items = [{"id": i.id, "category": i.category, "amount": i.amount, "description": i.description,
              "store_id": i.store_id, "method": i.method, "cash_account_id": i.cash_account_id,
              "income_date": i.income_date.isoformat()} for i in rows]
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@api.post("/admin/income")
async def create_income(payload: IncomeIn, db: AsyncSession = Depends(get_db), user: M.User = Depends(require_perm("manual_inc_exp"))):
    i = M.Income(**payload.model_dump(exclude_unset=True))
    i.created_by = user.user_id
    if not i.income_date:
        i.income_date = datetime.now(timezone.utc)
    db.add(i); await db.flush()
    if payload.cash_account_id:
        ca = (await db.execute(select(M.CashAccount).where(M.CashAccount.id == payload.cash_account_id))).scalar_one_or_none()
        if ca:
            ca.balance += payload.amount
            db.add(M.CashLedger(cash_account_id=ca.id, direction="in", amount=payload.amount,
                                source_kind="income", source_id=i.id,
                                notes=f"{i.category}: {i.description or ''}",
                                created_by=user.user_id))
    await db.commit(); await db.refresh(i)
    return {"id": i.id}


@api.delete("/admin/income/{iid}")
async def delete_income(iid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    i = (await db.execute(select(M.Income).where(M.Income.id == iid))).scalar_one_or_none()
    if not i:
        raise HTTPException(404, "Not found")
    await db.delete(i); await db.commit()
    return {"ok": True}


# ========== CASH ACCOUNTS ==========
class CashAccountIn(BaseModel):
    name: str
    kind: str = "cash"
    store_id: Optional[str] = None
    balance: float = 0.0
    active: bool = True


@api.get("/admin/cash-accounts")
async def list_cash_accounts(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    rows = (await db.execute(select(M.CashAccount).order_by(M.CashAccount.kind, M.CashAccount.name))).scalars().all()
    out = []
    for r in rows:
        store_name = ""
        if r.store_id:
            st = (await db.execute(select(M.Store).where(M.Store.id == r.store_id))).scalar_one_or_none()
            store_name = st.name if st else ""
        out.append({"id": r.id, "name": r.name, "kind": r.kind, "store_id": r.store_id,
                    "store_name": store_name, "balance": r.balance, "active": r.active})
    return out


@api.post("/admin/cash-accounts")
async def create_cash_account(payload: CashAccountIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    a = M.CashAccount(**payload.model_dump())
    db.add(a); await db.commit(); await db.refresh(a)
    return {"id": a.id}


@api.put("/admin/cash-accounts/{aid}")
async def update_cash_account(aid: str, payload: CashAccountIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    a = (await db.execute(select(M.CashAccount).where(M.CashAccount.id == aid))).scalar_one_or_none()
    if not a:
        raise HTTPException(404, "Not found")
    for k, v in payload.model_dump().items():
        setattr(a, k, v)
    await db.commit()
    return {"ok": True}


@api.delete("/admin/cash-accounts/{aid}")
async def delete_cash_account(aid: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    a = (await db.execute(select(M.CashAccount).where(M.CashAccount.id == aid))).scalar_one_or_none()
    if not a:
        raise HTTPException(404, "Not found")
    await db.delete(a); await db.commit()
    return {"ok": True}


@api.get("/admin/cash-ledger")
async def list_cash_ledger(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin),
                            cash_account_id: Optional[str] = None, limit: int = 200):
    base = select(M.CashLedger).order_by(desc(M.CashLedger.created_at)).limit(limit)
    if cash_account_id:
        base = base.where(M.CashLedger.cash_account_id == cash_account_id)
    rows = (await db.execute(base)).scalars().all()
    return [{"id": r.id, "cash_account_id": r.cash_account_id, "direction": r.direction,
             "amount": r.amount, "source_kind": r.source_kind, "source_id": r.source_id,
             "notes": r.notes, "created_at": r.created_at.isoformat()} for r in rows]


# ========== ACCOUNTING REPORTS (P&L) ==========
@api.get("/admin/reports/pnl")
async def pnl_report(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin),
                      from_date: Optional[str] = None, to_date: Optional[str] = None,
                      store_id: Optional[str] = None, group_by: str = "day"):
    """Profit & Loss with optional store filter and grouping (day or month)."""
    now = datetime.now(timezone.utc)
    start = datetime.fromisoformat(from_date) if from_date else now - timedelta(days=30)
    end = datetime.fromisoformat(to_date) if to_date else now
    if start.tzinfo is None: start = start.replace(tzinfo=timezone.utc)
    if end.tzinfo is None: end = end.replace(tzinfo=timezone.utc)
    # If the caller passed a date-only "to" (00:00 of the day) extend to end-of-day so that
    # entries logged later that same day still fall inside the window.
    if end.hour == 0 and end.minute == 0 and end.second == 0:
        end = end.replace(hour=23, minute=59, second=59, microsecond=999999)

    o_q = select(M.Order).where(and_(M.Order.created_at >= start, M.Order.created_at <= end, M.Order.payment_status == "paid"))
    if store_id:
        o_q = o_q.where(M.Order.store_id == store_id)
    orders = (await db.execute(o_q)).scalars().all()

    e_q = select(M.Expense).where(and_(M.Expense.expense_date >= start, M.Expense.expense_date <= end))
    if store_id:
        e_q = e_q.where(M.Expense.store_id == store_id)
    expenses = (await db.execute(e_q)).scalars().all()

    i_q = select(M.Income).where(and_(M.Income.income_date >= start, M.Income.income_date <= end))
    if store_id:
        i_q = i_q.where(M.Income.store_id == store_id)
    incomes = (await db.execute(i_q)).scalars().all()

    # Supplier payouts within window — they drain cash so count as expense.
    sp_q = select(M.SupplierPayment).where(and_(M.SupplierPayment.paid_date >= start, M.SupplierPayment.paid_date <= end))
    supplier_payments = (await db.execute(sp_q)).scalars().all()

    fmt = "%Y-%m" if group_by == "month" else "%Y-%m-%d"
    daily = {}
    def _b(d):
        return daily.setdefault(d, {"revenue": 0.0, "income": 0.0, "expense": 0.0})
    for o in orders:
        _b(o.created_at.strftime(fmt))["revenue"] += o.total
    for i in incomes:
        _b(i.income_date.strftime(fmt))["income"] += i.amount
    for e in expenses:
        _b(e.expense_date.strftime(fmt))["expense"] += e.amount
    # Supplier payouts roll into the expense bucket per their paid_date
    for sp in supplier_payments:
        _b(sp.paid_date.strftime(fmt))["expense"] += sp.amount
    series = []
    for k in sorted(daily.keys()):
        d = daily[k]
        series.append({"date": k, "revenue": round(d["revenue"], 2), "income": round(d["income"], 2),
                       "expense": round(d["expense"], 2),
                       "profit": round(d["revenue"] + d["income"] - d["expense"], 2)})
    total_rev = round(sum(o.total for o in orders), 2)
    total_inc = round(sum(i.amount for i in incomes), 2)
    total_exp = round(sum(e.amount for e in expenses) + sum(sp.amount for sp in supplier_payments), 2)
    total_supplier = round(sum(sp.amount for sp in supplier_payments), 2)
    # By outlet
    stores = {s.id: s.name for s in (await db.execute(select(M.Store))).scalars().all()}
    by_outlet = {}
    for o in orders:
        sid = o.store_id or "_unassigned"
        b = by_outlet.setdefault(sid, {"name": stores.get(sid, "Online/Unassigned"), "revenue": 0.0, "expense": 0.0, "income": 0.0})
        b["revenue"] += o.total
    for e in expenses:
        sid = e.store_id or "_unassigned"
        b = by_outlet.setdefault(sid, {"name": stores.get(sid, "Online/Unassigned"), "revenue": 0.0, "expense": 0.0, "income": 0.0})
        b["expense"] += e.amount
    for sp in supplier_payments:
        # Supplier payments aren't bound to a store; bucket as Unassigned.
        sid = "_unassigned"
        b = by_outlet.setdefault(sid, {"name": stores.get(sid, "Online/Unassigned"), "revenue": 0.0, "expense": 0.0, "income": 0.0})
        b["expense"] += sp.amount
    for i in incomes:
        sid = i.store_id or "_unassigned"
        b = by_outlet.setdefault(sid, {"name": stores.get(sid, "Online/Unassigned"), "revenue": 0.0, "expense": 0.0, "income": 0.0})
        b["income"] += i.amount
    by_outlet_arr = []
    for sid, v in by_outlet.items():
        by_outlet_arr.append({"store_id": sid, "store_name": v["name"], "revenue": round(v["revenue"], 2),
                               "income": round(v["income"], 2), "expense": round(v["expense"], 2),
                               "profit": round(v["revenue"] + v["income"] - v["expense"], 2)})
    return {
        "total_revenue": total_rev, "total_income": total_inc, "total_expense": total_exp,
        "supplier_payments": total_supplier,
        "profit": round(total_rev + total_inc - total_exp, 2),
        "series": series, "by_outlet": by_outlet_arr,
        "from": start.isoformat(), "to": end.isoformat(), "group_by": group_by,
    }


@api.get("/admin/reports/pnl/export")
async def pnl_export(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin),
                      from_date: Optional[str] = None, to_date: Optional[str] = None,
                      store_id: Optional[str] = None, group_by: str = "day"):
    """Excel xlsx export of P&L."""
    import io
    from openpyxl import Workbook
    data = await pnl_report(db, _=_, from_date=from_date, to_date=to_date, store_id=store_id, group_by=group_by)
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Summary"
    ws.append(["Period", data["from"], "to", data["to"]])
    ws.append([])
    ws.append(["Total Revenue", data["total_revenue"]])
    ws.append(["Total Income", data["total_income"]])
    ws.append(["Total Expense", data["total_expense"]])
    ws.append(["Net Profit", data["profit"]])
    ws.append([])
    ws.append([group_by.capitalize(), "Revenue", "Income", "Expense", "Profit"])
    for s in data["series"]:
        ws.append([s["date"], s["revenue"], s["income"], s["expense"], s["profit"]])
    ws2 = wb.create_sheet("By Outlet")
    ws2.append(["Outlet", "Revenue", "Income", "Expense", "Profit"])
    for o in data["by_outlet"]:
        ws2.append([o["store_name"], o["revenue"], o["income"], o["expense"], o["profit"]])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return FastResponse(content=buf.getvalue(),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f'attachment; filename="pnl_{group_by}.xlsx"'})


# ========== PUBLIC RECEIPT (for SMS link) ==========
@api.get("/receipt/{order_number}")
async def public_receipt(order_number: str, db: AsyncSession = Depends(get_db)):
    """Returns a public-friendly receipt JSON. Used by /receipt/{n} frontend page."""
    o = (await db.execute(select(M.Order).where(M.Order.order_number == order_number))).scalar_one_or_none()
    if not o:
        raise HTTPException(404, "Order not found")
    cs = (await db.execute(select(M.CompanySettings).where(M.CompanySettings.id == "default"))).scalar_one_or_none()
    items = (await db.execute(select(M.OrderItem).where(M.OrderItem.order_id == o.id))).scalars().all()
    return {
        "order_number": o.order_number, "status": o.status, "payment_status": o.payment_status,
        "payment_method": o.payment_method, "subtotal": o.subtotal, "discount": o.discount,
        "shipping": o.shipping, "total": o.total, "cash_tendered": o.cash_tendered, "cash_change": o.cash_change,
        "card_last4": o.card_last4, "customer_name": o.customer_name,
        "created_at": o.created_at.isoformat(),
        "items": [{"name": i.product_name, "variant": i.variant_label, "qty": i.quantity,
                   "unit": i.unit_price, "subtotal": i.subtotal} for i in items],
        "company": {"name": cs.company_name if cs else "Store", "address": cs.address if cs else None,
                    "phone": cs.phone if cs else None, "email": cs.email if cs else None,
                    "currency": cs.currency if cs else "LKR"},
    }


# ========== CSV IMPORT (Products + Inventory) ==========
class CsvImportIn(BaseModel):
    rows: List[dict]  # parsed CSV rows
    commit: bool = False  # if False, dry-run for preview


def _csv_clean(v):
    """Treat empty string / None as 'not provided'."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _csv_float(v):
    s = _csv_clean(v)
    if s is None: return None
    try: return float(s)
    except ValueError: return None


def _csv_int(v):
    s = _csv_clean(v)
    if s is None: return None
    try: return int(float(s))
    except ValueError: return None


def _csv_bool(v):
    s = _csv_clean(v)
    if s is None: return None
    return s.lower() in ("true", "1", "yes", "y", "t")


@api.post("/admin/import/products")
async def import_products(payload: CsvImportIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    """Bulk import products + first variant + initial inventory.
    Each row: name, sku, base_price, compare_price, cost_price, category(name), description,
              size, color, color_hex, stock, featured, status.
    Empty fields are skipped — only what's provided is updated.
    Existing product (matched by sku then name) updates the variant by size/color and adjusts inventory.
    """
    cats_by_name = {c.name.lower(): c for c in (await db.execute(select(M.Category))).scalars().all()}
    suppliers_by_name = {s.name.lower(): s for s in (await db.execute(select(M.Supplier))).scalars().all()}
    store = await _ensure_default_store(db)
    summary = {"created": 0, "updated": 0, "errors": []}
    preview = []
    for idx, raw in enumerate(payload.rows, start=1):
        try:
            name = _csv_clean(raw.get("name"))
            sku = _csv_clean(raw.get("sku"))
            if not name and not sku:
                raise ValueError("Row needs at least 'name' or 'sku'")
            base_price = _csv_float(raw.get("base_price"))
            compare_price = _csv_float(raw.get("compare_price"))
            cost_price = _csv_float(raw.get("cost_price"))
            category_name = _csv_clean(raw.get("category"))
            supplier_name = _csv_clean(raw.get("supplier"))
            description = _csv_clean(raw.get("description"))
            size = _csv_clean(raw.get("size"))
            color = _csv_clean(raw.get("color"))
            color_hex = _csv_clean(raw.get("color_hex"))
            stock = _csv_int(raw.get("stock"))
            featured = _csv_bool(raw.get("featured"))
            status_v = _csv_clean(raw.get("status"))
            cat = cats_by_name.get(category_name.lower()) if category_name else None
            sup = suppliers_by_name.get(supplier_name.lower()) if supplier_name else None
            existing = None
            if sku:
                existing = (await db.execute(select(M.Product).where(M.Product.sku == sku))).scalar_one_or_none()
            if not existing and name:
                existing = (await db.execute(select(M.Product).where(M.Product.name == name))).scalar_one_or_none()
            if existing:
                p = existing
                # Partial update: only override fields that were actually provided
                if name is not None: p.name = name
                if sku is not None: p.sku = sku
                if base_price is not None: p.base_price = base_price
                if compare_price is not None: p.compare_price = compare_price
                if cost_price is not None: p.cost_price = cost_price
                if cat: p.category_id = cat.id
                if sup: p.supplier_id = sup.id
                if description is not None: p.description = description
                if featured is not None: p.featured = featured
                if status_v is not None: p.status = status_v
                p.updated_at = datetime.now(timezone.utc)
                action = "updated"
            else:
                # New product needs a name
                if not name:
                    raise ValueError("New product requires 'name'")
                slug = slugify(name) + "-" + uuid.uuid4().hex[:4]
                p = M.Product(
                    name=name, slug=slug, description=description,
                    category_id=cat.id if cat else None,
                    supplier_id=sup.id if sup else None,
                    base_price=base_price if base_price is not None else 0.0,
                    compare_price=compare_price, cost_price=cost_price,
                    sku=sku, status=status_v or "active",
                    featured=bool(featured) if featured is not None else False,
                )
                db.add(p)
                await db.flush()
                action = "created"
            preview.append({"row": idx, "name": p.name, "sku": p.sku, "action": action,
                            "size": size, "color": color, "stock": stock if stock is not None else "—"})
            if action == "created": summary["created"] += 1
            else: summary["updated"] += 1
            if not payload.commit:
                continue
            # Variant + inventory
            if size or color:
                vq = select(M.Variant).where(and_(M.Variant.product_id == p.id, M.Variant.size == size, M.Variant.color == color))
                v = (await db.execute(vq)).scalar_one_or_none()
                if not v:
                    v = M.Variant(product_id=p.id, size=size, color=color, color_hex=color_hex)
                    db.add(v); await db.flush()
                else:
                    if color_hex: v.color_hex = color_hex
                if stock is not None:
                    inv = (await db.execute(select(M.Inventory).where(and_(
                        M.Inventory.variant_id == v.id, M.Inventory.store_id == store.id)))).scalar_one_or_none()
                    if inv:
                        inv.quantity = stock
                    else:
                        db.add(M.Inventory(variant_id=v.id, store_id=store.id, quantity=stock))
        except Exception as e:
            summary["errors"].append({"row": idx, "error": str(e)})
            continue
    if payload.commit:
        await db.commit()
    else:
        await db.rollback()
    return {"summary": summary, "preview": preview, "committed": payload.commit}


@api.get("/admin/import/products/template")
async def products_csv_template():
    csv = "name,sku,base_price,compare_price,cost_price,category,supplier,description,size,color,color_hex,stock,featured,status\n"
    csv += 'Sample Tee,TS-001,2500,3000,1500,Tees,,A soft cotton tee.,M,Black,#000000,12,false,active\n'
    csv += 'Sample Tee,TS-001,,,,,,,L,White,#ffffff,8,,\n'
    return FastResponse(content=csv, media_type="text/csv",
                        headers={"Content-Disposition": 'attachment; filename="products_template.csv"'})


# ========== DISCOUNTS (storefront promotions) ==========
class DiscountIn(BaseModel):
    name: str
    description: Optional[str] = None
    type: str = "percent"
    value: float = 0.0
    scope: str = "sitewide"
    scope_product_ids: Optional[List[str]] = None
    scope_category_ids: Optional[List[str]] = None
    show_badge_on_products: bool = True
    badge_label: Optional[str] = None
    badge_color: str = "#FF3B30"
    show_marquee: bool = True
    marquee_size: str = "sm"
    marquee_speed: str = "normal"
    marquee_bg: str = "#FF3B30"
    marquee_fg: str = "#FFFFFF"
    starts_at: Optional[datetime] = None
    ends_at: Optional[datetime] = None
    active: bool = True


def _discount_dict(d: M.Discount):
    return {
        "id": d.id, "name": d.name, "description": d.description,
        "type": d.type, "value": d.value, "scope": d.scope,
        "scope_product_ids": d.scope_product_ids or [],
        "scope_category_ids": d.scope_category_ids or [],
        "show_badge_on_products": d.show_badge_on_products,
        "badge_label": d.badge_label, "badge_color": d.badge_color,
        "show_marquee": d.show_marquee, "marquee_size": d.marquee_size,
        "marquee_speed": d.marquee_speed, "marquee_bg": d.marquee_bg, "marquee_fg": d.marquee_fg,
        "starts_at": d.starts_at.isoformat() if d.starts_at else None,
        "ends_at": d.ends_at.isoformat() if d.ends_at else None,
        "active": d.active,
        "created_at": d.created_at.isoformat() if d.created_at else None,
    }


@api.get("/admin/discounts")
async def list_discounts(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin),
                          q: Optional[str] = None, page: int = 1, page_size: int = 50):
    base = select(M.Discount)
    if q:
        base = base.where(or_(M.Discount.name.ilike(f"%{q}%"), M.Discount.description.ilike(f"%{q}%")))
    total = (await db.execute(select(func.count()).select_from(base.subquery()))).scalar_one()
    page_size = min(max(1, page_size), 100); page = max(1, page)
    rows = (await db.execute(base.order_by(desc(M.Discount.created_at)).offset((page-1)*page_size).limit(page_size))).scalars().all()
    return {"items": [_discount_dict(d) for d in rows], "total": total, "page": page, "page_size": page_size}


@api.post("/admin/discounts")
async def create_discount(payload: DiscountIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    d = M.Discount(**payload.model_dump())
    db.add(d); await db.commit(); await db.refresh(d)
    return _discount_dict(d)


@api.put("/admin/discounts/{did}")
async def update_discount(did: str, payload: DiscountIn, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    d = (await db.execute(select(M.Discount).where(M.Discount.id == did))).scalar_one_or_none()
    if not d:
        raise HTTPException(404, "Not found")
    for k, v in payload.model_dump().items():
        setattr(d, k, v)
    d.updated_at = datetime.now(timezone.utc)
    await db.commit()
    return _discount_dict(d)


@api.delete("/admin/discounts/{did}")
async def delete_discount(did: str, db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    d = (await db.execute(select(M.Discount).where(M.Discount.id == did))).scalar_one_or_none()
    if not d:
        raise HTTPException(404, "Not found")
    await db.delete(d); await db.commit()
    return {"ok": True}


@api.get("/discounts/active")
async def public_active_discounts(db: AsyncSession = Depends(get_db)):
    """Public list of currently-active discounts for marquee + product badges."""
    now = datetime.now(timezone.utc)
    rows = (await db.execute(select(M.Discount).where(M.Discount.active == True))).scalars().all()
    out = []
    for d in rows:
        if d.starts_at:
            sa = d.starts_at if d.starts_at.tzinfo else d.starts_at.replace(tzinfo=timezone.utc)
            if sa > now: continue
        if d.ends_at:
            ea = d.ends_at if d.ends_at.tzinfo else d.ends_at.replace(tzinfo=timezone.utc)
            if ea < now: continue
        out.append(_discount_dict(d))
    return out


# ========== CUSTOMER CSV / EXCEL EXPORT ==========
@api.get("/admin/customers/export.csv")
async def export_customers_csv(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    import io, csv as _csv
    rows = (await db.execute(select(M.Customer).order_by(desc(M.Customer.created_at)))).scalars().all()
    buf = io.StringIO()
    w = _csv.writer(buf)
    w.writerow(["name", "email", "phone", "address", "district", "city",
                "total_orders", "total_spent", "created_at"])
    for c in rows:
        w.writerow([c.name or "", c.email or "", c.phone or "", c.address or "",
                    c.district or "", c.city or "",
                    c.total_orders or 0, c.total_spent or 0,
                    c.created_at.isoformat() if c.created_at else ""])
    return FastResponse(content=buf.getvalue(), media_type="text/csv",
                        headers={"Content-Disposition": 'attachment; filename="customers.csv"'})


@api.get("/admin/customers/export.xlsx")
async def export_customers_xlsx(db: AsyncSession = Depends(get_db), _: M.User = Depends(require_admin)):
    import io
    from openpyxl import Workbook
    rows = (await db.execute(select(M.Customer).order_by(desc(M.Customer.created_at)))).scalars().all()
    wb = Workbook(); ws = wb.active; ws.title = "Customers"
    ws.append(["Name", "Email", "Phone", "Address", "District", "City",
               "Orders", "Total Spent", "Marketing Opt-in", "Joined"])
    for c in rows:
        ws.append([c.name or "", c.email or "", c.phone or "", c.address or "",
                   c.district or "", c.city or "",
                   c.total_orders or 0, c.total_spent or 0,
                   "yes" if c.marketing_opt_in else "no",
                   c.created_at.isoformat() if c.created_at else ""])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return FastResponse(content=buf.getvalue(),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": 'attachment; filename="customers.xlsx"'})


app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

